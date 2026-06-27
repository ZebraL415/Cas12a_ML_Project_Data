#!/usr/bin/env python3
"""Resolve EasyDesign round-1 questions and build baseline-ready v0 data.

The script is intentionally conservative:
- never writes to 01_raw;
- uses the combined source-data workbook as the authoritative table source;
- keeps measured labels, derived/augmented labels, and paper model predictions separate;
- writes paired Chinese and English markdown reports for this round.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


SOURCE_ID = "EasyDesign_2024"
PAPER_SHORT = "EasyDesign"
YEAR = "2024"
FULL_TITLE = "Deep learning enhancing guide RNA design for CRISPR/Cas12a-based diagnostics"

PDF_PATH = Path(
    "/Users/linzibo/Downloads/iMeta - 2024 - Huang - Deep learning enhancing guide RNA design for CRISPR Cas12a‐based diagnostics.pdf"
)
IMAGE_PATH = Path(
    "/var/folders/c9/xvbkv1hs5kz09g_fbjx8txd40000gn/T/codex-clipboard-e3e49006-b253-48c9-80e1-98d4742dde7b.png"
)

COMBINED_WORKBOOK = (
    Path("01_raw")
    / SOURCE_ID
    / "data"
    / "imt2214-sup-0002-tables1-9sourcedata (1).xlsx"
)

SUPPLEMENT_TABLE_DESCRIPTIONS = {
    "Table S1": "Information on the types of viral and bacterial species used to build crRNA libraries.",
    "Table S2": "Information on viral and bacterial nucleic acid templates used to build crRNA libraries.",
    "Table S3": "Training data datasets for model development generated based on CRISPR fluorescence reaction.",
    "Table S4": "Augment datasets for model development generated based on CRISPR fluorescence reaction.",
    "Table S5": "Test datasets for model development generated based on CRISPR fluorescence reaction.",
    "Table S6": "Model performance evaluation by k-fold cross-validation.",
    "Table S7": "The DNA templates of four pathogens (MPXV, EV71, CV-A16, and L. monocytogenes) were used in the experimental activity testing.",
    "Table S8": "The crRNAs of four pathogens (MPXV, EV71, CV-A16, and L. monocytogenes) were used in the experimental activity testing.",
    "Table S9": "Optimal HPV crRNAs and RPA primers generated from the one-stop web platform of EasyDesign.",
}

CATALOG_COLUMNS = [
    "source_id",
    "paper_short",
    "year",
    "full_title",
    "path_type",
    "source_type",
    "raw_path",
    "file_names",
    "data_access_status",
    "record_unit",
    "label_type",
    "label_status",
    "sample_size_estimated",
    "sequence_fields_detected",
    "usable_for_training",
    "usable_for_extension",
    "priority",
    "main_risk",
    "notes",
]

SHEET_INDEX_COLUMNS = [
    "source_id",
    "file_name",
    "file_path",
    "sheet_name",
    "n_rows",
    "n_cols",
    "first_columns",
    "example_values",
    "guessed_content",
    "record_unit_guess",
    "has_crRNA_sequence",
    "has_target_sequence",
    "has_PAM",
    "has_label",
    "label_raw_candidates",
    "label_status_guess",
    "path_type_guess",
    "priority",
    "recommended_action",
    "extracted_csv_path",
    "notes",
    "evidence_source",
    "title_or_caption",
]

LABEL_COLUMNS = [
    "label_raw_name",
    "normalized_label",
    "path_type",
    "biological_meaning_cn",
    "biological_meaning_en",
    "assay_readout",
    "label_status",
    "trainable_as_primary_label",
    "standard_unit",
    "transform_needed",
    "notes",
]

MINIMAL_COLUMNS = [
    "dataset_id",
    "source_id",
    "paper_short",
    "year",
    "source_table_id",
    "source_file",
    "source_sheet",
    "record_id",
    "record_id_original",
    "paper_split",
    "baseline_split",
    "data_role",
    "organism",
    "target_name",
    "cas_type",
    "crRNA_sequence",
    "target_sequence",
    "target_context_sequence",
    "target_sequence_role",
    "pam",
    "pam_inference_status",
    "guide_target_hamming_dist_raw",
    "guide_target_hamming_dist_computed",
    "target_window_start_0based",
    "target_window_method",
    "label_raw_name",
    "label_raw_value",
    "label_normalized",
    "label_scale_group",
    "label_status",
    "label_role",
    "is_measured",
    "paper_prediction_CNND",
    "paper_prediction_CNN12a",
    "paper_prediction_CNN12ae",
    "paper_prediction_TransformerD",
    "paper_prediction_Transformer12a",
    "paper_prediction_Transformer12ae",
    "notes",
]

FEATURE_COLUMNS = [
    "dataset_id",
    "source_id",
    "source_table_id",
    "record_id",
    "paper_split",
    "baseline_split",
    "data_role",
    "crRNA_sequence",
    "target_sequence",
    "target_context_sequence",
    "pam",
    "label_raw_name",
    "label_raw_value",
    "label_normalized",
    "label_scale_group",
    "crRNA_length",
    "target_length",
    "target_context_length",
    "crRNA_GC_content",
    "target_GC_content",
    "target_context_GC_content",
    "guide_target_hamming_dist_computed",
    "guide_target_hamming_dist_raw",
    "has_valid_DNA_alphabet",
    "contains_ambiguous_base",
    "pam_type",
    "label_is_primary_baseline",
    "notes",
]


@dataclass
class Paths:
    root: Path
    catalog: Path
    notes: Path
    cleaned: Path
    candidate: Path
    diagnostic_extracted: Path
    scripts: Path
    raw_combined: Path


def project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def clean_text(value: Any, max_len: int = 500) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def ensure_paths(root: Path) -> Paths:
    paths = Paths(
        root=root,
        catalog=root / "00_data_catalog",
        notes=root / "99_notes",
        cleaned=root / "03_cleaned_minimal",
        candidate=root / "04_candidate_ml_dataset",
        diagnostic_extracted=root / "02_extracted_tables" / "diagnostic_activity",
        scripts=root / "scripts",
        raw_combined=root / COMBINED_WORKBOOK,
    )
    if not paths.raw_combined.exists():
        raise FileNotFoundError(f"Missing combined workbook: {paths.raw_combined}")
    for directory in [
        paths.catalog,
        paths.notes,
        paths.cleaned,
        paths.candidate,
        paths.diagnostic_extracted,
        paths.scripts,
    ]:
        directory.mkdir(parents=True, exist_ok=True)
    return paths


def backup(path: Path, timestamp: str) -> Path | None:
    if not path.exists():
        return None
    backup_path = path.with_name(f"{path.stem}_{timestamp}_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def safe_rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def read_xlsx(path: Path, columns: list[str]) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=columns)
    df = pd.read_excel(path, dtype=str).fillna("")
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns]


def style_xlsx(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col_idx, col_cells in enumerate(ws.columns, 1):
            width = 12
            for cell in list(col_cells)[:200]:
                width = max(width, min(len(clean_text(cell.value, 160)) + 2, 60))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
    wb.save(path)


def write_xlsx(df: pd.DataFrame, path: Path, columns: list[str], timestamp: str) -> None:
    backup(path, timestamp)
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    out = out[columns].fillna("")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=path.stem[:31])
    style_xlsx(path)


def norm_seq(value: Any) -> str:
    return re.sub(r"[^A-Za-z]", "", clean_text(value, max_len=10000)).upper()


def valid_dna_or_rna(seq: str) -> bool:
    return bool(seq) and bool(re.fullmatch(r"[ACGTU]+", seq))


def contains_ambiguous(seq: str) -> bool:
    return bool(seq) and not bool(re.fullmatch(r"[ACGTU]+", seq))


def gc_content(seq: str) -> float | None:
    seq = norm_seq(seq).replace("U", "T")
    if not seq:
        return None
    return round((seq.count("G") + seq.count("C")) / len(seq), 6)


def hamming(a: str, b: str) -> int | None:
    a = norm_seq(a)
    b = norm_seq(b)
    if not a or not b:
        return None
    return sum(x != y for x, y in zip(a, b)) + abs(len(a) - len(b))


def best_window(target_context: str, guide: str) -> tuple[str, int | None, int | None]:
    target_context = norm_seq(target_context)
    guide = norm_seq(guide)
    if not target_context or not guide or len(target_context) < len(guide):
        return "", None, None
    candidates = []
    for start in range(0, len(target_context) - len(guide) + 1):
        window = target_context[start : start + len(guide)]
        candidates.append((hamming(window, guide), start, window))
    best = min(candidates, key=lambda item: (999 if item[0] is None else item[0], item[1]))
    return best[2], best[1], best[0]


def target_window_from_context(target_context: str, guide: str) -> tuple[str, int | None, int | None, str]:
    """Return the 25-nt target window from a 45-nt context.

    The paper describes a 21-nt sequence downstream of the PAM with 10-nt
    flanks on both sides, producing a 45-nt target context. Therefore, when
    the expected lengths are present, the source-defined target window starts
    at 0-based position 10. The best-match search is only a fallback.
    """
    target_context = norm_seq(target_context)
    guide = norm_seq(guide)
    if len(target_context) == 45 and len(guide) == 25:
        window = target_context[10:35]
        return window, 10, hamming(window, guide), "source_defined_10nt_flank_window"
    window, start, dist = best_window(target_context, guide)
    return window, start, dist, "best_match_fallback"


def infer_pam(*seqs: str) -> tuple[str, str]:
    for seq in seqs:
        seq = norm_seq(seq).replace("U", "T")
        if len(seq) >= 4 and re.fullmatch(r"TTT[ACGT]", seq[:4]):
            return seq[:4], "inferred_from_5prime_TTTN_prefix_not_raw_PAM_column"
    return "", "not_detected_as_separate_column"


def hash_bucket(value: str) -> int:
    return int(hashlib.sha1(value.encode("utf-8")).hexdigest()[:8], 16) % 100


def baseline_split_for_train_target(target_sequence: str) -> str:
    return "baseline_train" if hash_bucket(norm_seq(target_sequence)) < 80 else "baseline_validation"


def numeric(value: Any) -> float | None:
    try:
        if pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def build_primary_dataset(paths: Paths) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    raw_path = paths.raw_combined
    train = pd.read_excel(raw_path, sheet_name="Table S3", header=1)
    augment = pd.read_excel(raw_path, sheet_name="Table S4", header=1)
    test = pd.read_excel(raw_path, sheet_name="Table S5", header=1)

    primary_rows: list[dict[str, Any]] = []
    for _, row in train.iterrows():
        guide = norm_seq(row["guide_seq"])
        target = norm_seq(row["target_at_guide"])
        pam, pam_status = infer_pam(target, guide)
        record_id = f"EasyDesign_2024_TableS3_{int(row['No.']):05d}"
        primary_rows.append(
            {
                "dataset_id": "EasyDesign_2024_diagnostic_activity_v0",
                "source_id": SOURCE_ID,
                "paper_short": PAPER_SHORT,
                "year": YEAR,
                "source_table_id": "EasyDesign_2024_TableS3_training",
                "source_file": safe_rel(raw_path, paths.root),
                "source_sheet": "Table S3",
                "record_id": record_id,
                "record_id_original": int(row["No."]),
                "paper_split": "paper_training",
                "baseline_split": baseline_split_for_train_target(target),
                "data_role": "primary_baseline_internal",
                "organism": clean_text(row.get("type1", "")),
                "target_name": clean_text(row.get("type2", "")),
                "cas_type": "Cas12a",
                "crRNA_sequence": guide,
                "target_sequence": target,
                "target_context_sequence": "",
                "target_sequence_role": "target_at_guide_25nt_with_TTTN_prefix",
                "pam": pam,
                "pam_inference_status": pam_status,
                "guide_target_hamming_dist_raw": numeric(row.get("guide_target_hamming_dist")),
                "guide_target_hamming_dist_computed": hamming(guide, target),
                "target_window_start_0based": "",
                "target_window_method": "source_provided_25nt_target_at_guide",
                "label_raw_name": "30 min",
                "label_raw_value": numeric(row["30 min"]),
                "label_normalized": numeric(row["30 min"]),
                "label_scale_group": "table_s3_log_or_transformed_30min_activity",
                "label_status": "measured",
                "label_role": "primary_baseline_label_within_TableS3_scale",
                "is_measured": "yes",
                "paper_prediction_CNND": "",
                "paper_prediction_CNN12a": "",
                "paper_prediction_CNN12ae": "",
                "paper_prediction_TransformerD": "",
                "paper_prediction_Transformer12a": "",
                "paper_prediction_Transformer12ae": "",
                "notes": "Primary internal baseline label. PDF confirms 30-min fluorescence was selected as activity indicator; numeric scale should not be mixed with Table S5 without transform confirmation.",
            }
        )

    for _, row in test.iterrows():
        guide = norm_seq(row["crRNA"])
        context = norm_seq(row["DNA"])
        target, start, best_ham, window_method = target_window_from_context(context, guide)
        pam, pam_status = infer_pam(target, guide, context)
        record_id = f"EasyDesign_2024_TableS5_{int(row['No.']):05d}"
        primary_rows.append(
            {
                "dataset_id": "EasyDesign_2024_diagnostic_activity_v0",
                "source_id": SOURCE_ID,
                "paper_short": PAPER_SHORT,
                "year": YEAR,
                "source_table_id": "EasyDesign_2024_TableS5_test",
                "source_file": safe_rel(raw_path, paths.root),
                "source_sheet": "Table S5",
                "record_id": record_id,
                "record_id_original": int(row["No."]),
                "paper_split": "paper_test",
                "baseline_split": "external_test_scale_unconfirmed",
                "data_role": "external_paper_test_holdout",
                "organism": clean_text(row.get("type", "")),
                "target_name": "",
                "cas_type": "Cas12a",
                "crRNA_sequence": guide,
                "target_sequence": target,
                "target_context_sequence": context,
                "target_sequence_role": "best_25nt_window_from_45nt_DNA_context",
                "pam": pam,
                "pam_inference_status": pam_status,
                "guide_target_hamming_dist_raw": "",
                "guide_target_hamming_dist_computed": best_ham,
                "target_window_start_0based": start,
                "target_window_method": window_method,
                "label_raw_name": "true value",
                "label_raw_value": numeric(row["true value"]),
                "label_normalized": numeric(row["true value"]),
                "label_scale_group": "table_s5_true_activity_scale_unconfirmed_against_TableS3",
                "label_status": "measured",
                "label_role": "external_test_label_scale_unconfirmed",
                "is_measured": "yes",
                "paper_prediction_CNND": numeric(row.get("CNND")),
                "paper_prediction_CNN12a": numeric(row.get("CNN12a")),
                "paper_prediction_CNN12ae": numeric(row.get("CNN12ae")),
                "paper_prediction_TransformerD": numeric(row.get("TransformerD")),
                "paper_prediction_Transformer12a": numeric(row.get("Transformer12a")),
                "paper_prediction_Transformer12ae": numeric(row.get("Transformer12ae")),
                "notes": f"Measured paper test true value. Paper model prediction columns are retained separately and must not be used as labels. target_window_method={window_method}.",
            }
        )

    augment_rows: list[dict[str, Any]] = []
    for _, row in augment.iterrows():
        guide = norm_seq(row["guide_seq"])
        target = norm_seq(row["target_at_guide"])
        pam, pam_status = infer_pam(target, guide)
        record_id = f"EasyDesign_2024_TableS4_{int(row['No.']):05d}"
        augment_rows.append(
            {
                "dataset_id": "EasyDesign_2024_diagnostic_activity_augmented_optional_v0",
                "source_id": SOURCE_ID,
                "paper_short": PAPER_SHORT,
                "year": YEAR,
                "source_table_id": "EasyDesign_2024_TableS4_augment",
                "source_file": safe_rel(raw_path, paths.root),
                "source_sheet": "Table S4",
                "record_id": record_id,
                "record_id_original": int(row["No."]),
                "paper_split": "paper_augment",
                "baseline_split": "optional_augmentation_not_in_primary_baseline",
                "data_role": "optional_augmentation",
                "organism": "",
                "target_name": "",
                "cas_type": "Cas12a",
                "crRNA_sequence": guide,
                "target_sequence": target,
                "target_context_sequence": "",
                "target_sequence_role": "target_at_guide_25nt_with_TTTN_prefix",
                "pam": pam,
                "pam_inference_status": pam_status,
                "guide_target_hamming_dist_raw": "",
                "guide_target_hamming_dist_computed": hamming(guide, target),
                "target_window_start_0based": "",
                "target_window_method": "source_provided_25nt_target_at_guide",
                "label_raw_name": "out_logk_measurement",
                "label_raw_value": numeric(row["out_logk_measurement"]),
                "label_normalized": numeric(row["out_logk_measurement"]),
                "label_scale_group": "table_s4_augmented_out_logk_measurement",
                "label_status": "measured",
                "label_role": "optional_derived_augmentation_label",
                "is_measured": "yes",
                "paper_prediction_CNND": "",
                "paper_prediction_CNN12a": "",
                "paper_prediction_CNN12ae": "",
                "paper_prediction_TransformerD": "",
                "paper_prediction_Transformer12a": "",
                "paper_prediction_Transformer12ae": "",
                "notes": "Optional augmented/derived label from Table S4; do not mix into first primary baseline without an explicit augmentation setting.",
            }
        )

    primary_df = pd.DataFrame(primary_rows, columns=MINIMAL_COLUMNS)
    augment_df = pd.DataFrame(augment_rows, columns=MINIMAL_COLUMNS)
    stats = {
        "table_s3_rows": int(len(train)),
        "table_s4_rows": int(len(augment)),
        "table_s5_rows": int(len(test)),
        "primary_rows": int(len(primary_df)),
        "augment_rows": int(len(augment_df)),
        "baseline_train_rows": int((primary_df["baseline_split"] == "baseline_train").sum()),
        "baseline_validation_rows": int((primary_df["baseline_split"] == "baseline_validation").sum()),
        "external_test_rows": int((primary_df["baseline_split"] == "external_test_scale_unconfirmed").sum()),
        "table_s3_unique_target_sequences": int(train["target_at_guide"].nunique()),
        "table_s5_best_window_start_counts": {
            str(k): int(v)
            for k, v in primary_df.loc[
                primary_df["source_table_id"] == "EasyDesign_2024_TableS5_test",
                "target_window_start_0based",
            ]
            .value_counts()
            .to_dict()
            .items()
        },
        "table_s5_window_method_counts": {
            str(k): int(v)
            for k, v in primary_df.loc[
                primary_df["source_table_id"] == "EasyDesign_2024_TableS5_test",
                "target_window_method",
            ]
            .value_counts()
            .to_dict()
            .items()
        },
        "table_s5_context_length_counts": {
            str(k): int(v)
            for k, v in primary_df.loc[
                primary_df["source_table_id"] == "EasyDesign_2024_TableS5_test",
                "target_context_sequence",
            ]
            .str.len()
            .value_counts()
            .to_dict()
            .items()
        },
    }
    return primary_df, augment_df, stats


def build_feature_table(primary_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in primary_df.iterrows():
        guide = row["crRNA_sequence"]
        target = row["target_sequence"]
        context = row["target_context_sequence"]
        alphabet_seq = guide + target + context
        rows.append(
            {
                "dataset_id": row["dataset_id"],
                "source_id": row["source_id"],
                "source_table_id": row["source_table_id"],
                "record_id": row["record_id"],
                "paper_split": row["paper_split"],
                "baseline_split": row["baseline_split"],
                "data_role": row["data_role"],
                "crRNA_sequence": guide,
                "target_sequence": target,
                "target_context_sequence": context,
                "pam": row["pam"],
                "label_raw_name": row["label_raw_name"],
                "label_raw_value": row["label_raw_value"],
                "label_normalized": row["label_normalized"],
                "label_scale_group": row["label_scale_group"],
                "crRNA_length": len(norm_seq(guide)),
                "target_length": len(norm_seq(target)),
                "target_context_length": len(norm_seq(context)) if context else 0,
                "crRNA_GC_content": gc_content(guide),
                "target_GC_content": gc_content(target),
                "target_context_GC_content": gc_content(context) if context else "",
                "guide_target_hamming_dist_computed": row["guide_target_hamming_dist_computed"],
                "guide_target_hamming_dist_raw": row["guide_target_hamming_dist_raw"],
                "has_valid_DNA_alphabet": "yes" if valid_dna_or_rna(alphabet_seq) else "no",
                "contains_ambiguous_base": "yes" if contains_ambiguous(alphabet_seq) else "no",
                "pam_type": "TTTN_inferred" if row["pam"] else "not_detected",
                "label_is_primary_baseline": "yes"
                if row["baseline_split"] in {"baseline_train", "baseline_validation"}
                else "no",
                "notes": row["notes"],
            }
        )
    return pd.DataFrame(rows, columns=FEATURE_COLUMNS)


def write_csv_with_backup(df: pd.DataFrame, path: Path, timestamp: str) -> None:
    backup(path, timestamp)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def export_authoritative_raw_csvs(paths: Paths, timestamp: str) -> dict[str, str]:
    outputs = {}
    sheets = {
        "Table S3": ("EasyDesign_2024_TableS3_diagnostic_activity_raw.csv", 1),
        "Table S4": ("EasyDesign_2024_TableS4_diagnostic_activity_raw.csv", 1),
        "Table S5": ("EasyDesign_2024_TableS5_diagnostic_activity_raw.csv", 1),
    }
    for sheet_name, (file_name, header_row) in sheets.items():
        df = pd.read_excel(paths.raw_combined, sheet_name=sheet_name, header=header_row)
        out_path = paths.diagnostic_extracted / file_name
        write_csv_with_backup(df, out_path, timestamp)
        outputs[sheet_name] = safe_rel(out_path, paths.root)
    return outputs


def extract_pdf_facts() -> list[dict[str, str]]:
    if not PDF_PATH.exists():
        return [
            {
                "source": str(PDF_PATH),
                "page": "",
                "fact_id": "pdf_missing",
                "fact_en": "PDF file was not available at runtime.",
                "fact_zh": "运行时未找到 PDF 文件。",
            }
        ]
    reader = PdfReader(str(PDF_PATH))
    pages = [page.extract_text() or "" for page in reader.pages]

    def find_page(term: str) -> str:
        for idx, text in enumerate(pages, 1):
            if term.lower() in text.lower():
                return str(idx)
        return ""

    return [
        {
            "source": str(PDF_PATH),
            "page": find_page("11,496 experimentally validated"),
            "fact_id": "experimentally_validated_cases",
            "fact_en": "The paper reports 11,496 experimentally validated Cas12a-based detection cases.",
            "fact_zh": "论文报告共有 11,496 个经实验验证的 Cas12a 检测案例。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("fluorescence value at 30 min was selected"),
            "fact_id": "30min_primary_indicator",
            "fact_en": "The methods state that the fluorescence value at 30 minutes was selected as the activity indicator.",
            "fact_zh": "方法部分说明选择 30 分钟荧光值作为活性评估指标。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("subtracted the mean value of the negative controls"),
            "fact_id": "control_normalization",
            "fact_en": "Plate readings were normalized using negative controls and positive controls in the same plate.",
            "fact_zh": "论文说明使用同板阴性对照和阳性对照对读板数值进行归一化。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("two label values were used"),
            "fact_id": "20min_timefolding",
            "fact_en": "Data augmentation used two label values for each pair: the 30-minute fluorescence and a 20-minute readout normalized to 30 minutes.",
            "fact_zh": "数据增强对每个配对使用两个标签值：30 分钟荧光值，以及由 20 分钟读数归一到 30 分钟的值。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("containing 31,993 guide"),
            "fact_id": "augment_size",
            "fact_en": "The augmented dataset is described as containing 31,993 guide-to-target pairs.",
            "fact_zh": "论文描述增强数据集包含 31,993 个 guide-to-target 配对。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("training set encompassed 20 distinct classes"),
            "fact_id": "method_split_10634_1358",
            "fact_en": "The methods describe 10,634 training pairs and 1,358 testing pairs.",
            "fact_zh": "方法部分描述训练集包含 10,634 个配对，测试集包含 1,358 个配对。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("resulting in a 45"),
            "fact_id": "target_context_45nt",
            "fact_en": "The methods describe a 21-nt sequence downstream of the PAM plus 10-nt flanks, producing a 45-nt target context.",
            "fact_zh": "方法部分描述提取 PAM 下游 21 nt，并在两端各扩展 10 nt，形成 45 nt target context。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("TTTN"),
            "fact_id": "pam_tttn",
            "fact_en": "The paper discusses a TTTN PAM for the Cas12a guide-target sequence pairs.",
            "fact_zh": "论文讨论了 Cas12a guide-target 配对中的 TTTN PAM。",
        },
        {
            "source": str(PDF_PATH),
            "page": find_page("Table S3: Training data"),
            "fact_id": "supplement_table_descriptions",
            "fact_en": "The supporting information list describes Tables S3-S5 as model-development datasets based on CRISPR fluorescence reaction.",
            "fact_zh": "补充资料列表说明 Table S3-S5 是基于 CRISPR fluorescence reaction 的模型开发数据集。",
        },
    ]


def image_facts() -> list[dict[str, str]]:
    return [
        {
            "source": str(IMAGE_PATH),
            "fact_id": table_id,
            "fact_en": desc,
            "fact_zh": {
                "Table S1": "用于构建 crRNA 文库的病毒和细菌物种类型信息。",
                "Table S2": "用于构建 crRNA 文库的病毒和细菌核酸模板信息。",
                "Table S3": "基于 CRISPR 荧光反应生成的模型开发训练数据集。",
                "Table S4": "基于 CRISPR 荧光反应生成的模型开发增强数据集。",
                "Table S5": "基于 CRISPR 荧光反应生成的模型开发测试数据集。",
                "Table S6": "k-fold 交叉验证的模型性能评估。",
                "Table S7": "四种病原体的 DNA 模板，用于实验活性测试。",
                "Table S8": "四种病原体的 crRNAs，用于实验活性测试。",
                "Table S9": "由 EasyDesign 一站式网页平台生成的最优 HPV crRNAs 和 RPA primers。",
            }[table_id],
        }
        for table_id, desc in SUPPLEMENT_TABLE_DESCRIPTIONS.items()
    ]


def update_catalogs(
    paths: Paths,
    primary_df: pd.DataFrame,
    augment_df: pd.DataFrame,
    raw_exports: dict[str, str],
    timestamp: str,
) -> list[str]:
    outputs = []

    master_path = paths.catalog / "master_data_catalog.xlsx"
    master = read_xlsx(master_path, CATALOG_COLUMNS)
    master = master[master["source_id"] != SOURCE_ID]
    source_row = {
        "source_id": SOURCE_ID,
        "paper_short": PAPER_SHORT,
        "year": YEAR,
        "full_title": FULL_TITLE,
        "path_type": "diagnostic_activity",
        "source_type": "Original paper PDF plus combined supplementary source-data workbook; standalone tables treated as duplicates.",
        "raw_path": f"01_raw/{SOURCE_ID}",
        "file_names": safe_rel(paths.raw_combined, paths.root),
        "data_access_status": "local raw files available; PDF evidence supplied; 01_raw not modified",
        "record_unit": "crRNA-target pair for Table S3/S4/S5; metadata/figure source data for other tables",
        "label_type": "30-min fluorescence/activity; 20-min-to-30-min derived activity; out_logk augmented activity; true-value test activity; paper model predictions kept separate",
        "label_status": "measured",
        "sample_size_estimated": (
            f"primary baseline rows={len(primary_df)}; Table S3 internal train/validation={len(primary_df[primary_df['source_table_id']=='EasyDesign_2024_TableS3_training'])}; "
            f"Table S5 external test={len(primary_df[primary_df['source_table_id']=='EasyDesign_2024_TableS5_test'])}; optional Table S4 augmentation={len(augment_df)}"
        ),
        "sequence_fields_detected": "guide_seq/crRNA; target_at_guide; DNA 45-nt context; inferred TTTN PAM prefix",
        "usable_for_training": "yes",
        "usable_for_extension": "maybe",
        "priority": "high",
        "main_risk": "Table S3 30-min values and Table S5 true values are both fluorescence-derived but not proven to share a numeric scale; use Table S3 internal split first.",
        "notes": "Round2 resolved evidence from PDF and supplement screenshot. Combined source-data workbook is the authoritative source. Table S4 is optional augmentation, not primary baseline.",
    }
    master = pd.concat([master, pd.DataFrame([source_row])], ignore_index=True)
    write_xlsx(master, master_path, CATALOG_COLUMNS, timestamp)
    outputs.append(safe_rel(master_path, paths.root))

    sheet_path = paths.catalog / "source_sheet_index.xlsx"
    sheet = read_xlsx(sheet_path, SHEET_INDEX_COLUMNS)
    for idx, row in sheet.iterrows():
        if row.get("source_id") != SOURCE_ID:
            continue
        sheet_name = row.get("sheet_name")
        file_name = row.get("file_name")
        note = clean_text(row.get("notes", ""))
        if file_name == paths.raw_combined.name and sheet_name in SUPPLEMENT_TABLE_DESCRIPTIONS:
            sheet.at[idx, "title_or_caption"] = SUPPLEMENT_TABLE_DESCRIPTIONS[sheet_name]
            sheet.at[idx, "evidence_source"] = "PDF supporting-information list and supplied supplement screenshot"
        if sheet_name == "Table S2":
            sheet.at[idx, "path_type_guess"] = "metadata_only"
            sheet.at[idx, "record_unit_guess"] = "genomic target site"
            sheet.at[idx, "recommended_action"] = "keep_as_metadata"
            sheet.at[idx, "notes"] = "Round2: supplement description identifies this as nucleic acid template metadata for crRNA library construction."
        elif sheet_name == "Table S3":
            sheet.at[idx, "path_type_guess"] = "diagnostic_activity"
            sheet.at[idx, "record_unit_guess"] = "crRNA-target pair"
            sheet.at[idx, "label_status_guess"] = "measured"
            sheet.at[idx, "priority"] = "high"
            sheet.at[idx, "recommended_action"] = "extract_to_csv"
            sheet.at[idx, "extracted_csv_path"] = raw_exports.get("Table S3", "")
            sheet.at[idx, "notes"] = "Round2: authoritative training table. Use 30 min as primary internal-baseline label; keep 20 min normalized as auxiliary derived label."
        elif sheet_name == "Table S4":
            sheet.at[idx, "path_type_guess"] = "diagnostic_activity"
            sheet.at[idx, "record_unit_guess"] = "crRNA-target pair"
            sheet.at[idx, "label_status_guess"] = "measured"
            sheet.at[idx, "priority"] = "high"
            sheet.at[idx, "recommended_action"] = "extract_to_csv"
            sheet.at[idx, "extracted_csv_path"] = raw_exports.get("Table S4", "")
            sheet.at[idx, "notes"] = "Round2: authoritative optional augmentation table. Do not mix into primary baseline unless augmentation is explicitly enabled."
        elif sheet_name == "Table S5":
            sheet.at[idx, "path_type_guess"] = "diagnostic_activity"
            sheet.at[idx, "record_unit_guess"] = "crRNA-target pair"
            sheet.at[idx, "label_status_guess"] = "measured"
            sheet.at[idx, "priority"] = "high"
            sheet.at[idx, "recommended_action"] = "extract_to_csv"
            sheet.at[idx, "extracted_csv_path"] = raw_exports.get("Table S5", "")
            sheet.at[idx, "notes"] = "Round2: authoritative paper test table. true value is measured; CNND/CNN/Transformer columns are paper model predictions, not labels."
        elif sheet_name in {"Table S7", "Table S8"}:
            sheet.at[idx, "path_type_guess"] = "metadata_only"
            sheet.at[idx, "record_unit_guess"] = "experimental validation metadata"
            sheet.at[idx, "priority"] = "medium"
            sheet.at[idx, "recommended_action"] = "keep_as_metadata"
            sheet.at[idx, "notes"] = "Round2: supplement description says these DNA templates/crRNAs were used in experimental activity testing; no row-level label in this sheet."
        elif sheet_name == "Fig.S3":
            sheet.at[idx, "path_type_guess"] = "metadata_only"
            sheet.at[idx, "record_unit_guess"] = "figure source data"
            sheet.at[idx, "label_status_guess"] = "metadata"
            sheet.at[idx, "priority"] = "low"
            sheet.at[idx, "recommended_action"] = "keep_as_metadata"
            sheet.at[idx, "notes"] = "Round2: aggregate/correlation figure source data; not a tidy crRNA-target training table."
        elif sheet_name in {"Fig.3", "Fig.4", "Fig.S2", "Fig.S7", "Fig.S8"}:
            sheet.at[idx, "path_type_guess"] = "diagnostic_activity"
            sheet.at[idx, "record_unit_guess"] = "figure source data"
            sheet.at[idx, "priority"] = "medium"
            sheet.at[idx, "recommended_action"] = "inspect_manually"
            sheet.at[idx, "notes"] = (note + " Round2: fluorescence figure source data; do not use as primary baseline rows.").strip()
    write_xlsx(sheet, sheet_path, SHEET_INDEX_COLUMNS, timestamp)
    outputs.append(safe_rel(sheet_path, paths.root))

    label_path = paths.catalog / "label_dictionary.xlsx"
    label = read_xlsx(label_path, LABEL_COLUMNS)
    label = label[~label["notes"].str.contains("source_id=EasyDesign_2024", na=False)]
    rows = [
        {
            "label_raw_name": "30 min",
            "normalized_label": "diagnostic_activity_30min_primary",
            "path_type": "diagnostic_activity",
            "biological_meaning_cn": "30 分钟 Cas12a 诊断荧光/活性指标；本轮作为 Table S3 内部 baseline 主标签。",
            "biological_meaning_en": "30-minute Cas12a diagnostic fluorescence/activity indicator; used as the Table S3 internal-baseline primary label in this round.",
            "assay_readout": "control-normalized fluorescence/activity, numeric scale as Table S3",
            "label_status": "measured",
            "trainable_as_primary_label": "yes",
            "standard_unit": "Table S3 numeric scale; exact transform to Table S5 not confirmed",
            "transform_needed": "no for Table S3 internal baseline; yes before cross-scale use",
            "notes": "source_id=EasyDesign_2024; PDF methods select 30-min fluorescence as activity indicator.",
        },
        {
            "label_raw_name": "20 min normalized",
            "normalized_label": "diagnostic_activity_20min_to_30min_auxiliary",
            "path_type": "diagnostic_activity",
            "biological_meaning_cn": "由 20 分钟读数归一到 30 分钟的派生活性标签。",
            "biological_meaning_en": "Derived activity label from a 20-minute readout normalized to 30 minutes.",
            "assay_readout": "time-folded normalized fluorescence/activity",
            "label_status": "measured",
            "trainable_as_primary_label": "no",
            "standard_unit": "derived Table S3 auxiliary scale",
            "transform_needed": "already derived; keep separate from 30 min unless augmentation is enabled",
            "notes": "source_id=EasyDesign_2024; PDF describes two labels per pair for time-folding/data augmentation.",
        },
        {
            "label_raw_name": "out_logk_measurement",
            "normalized_label": "diagnostic_activity_augmented_out_logk",
            "path_type": "diagnostic_activity",
            "biological_meaning_cn": "Table S4 增强数据中的派生/增强活性测量值。",
            "biological_meaning_en": "Derived/augmented activity measurement in the Table S4 augmented dataset.",
            "assay_readout": "augmented activity measurement derived from CRISPR fluorescence reaction",
            "label_status": "measured",
            "trainable_as_primary_label": "no",
            "standard_unit": "Table S4 augmented scale",
            "transform_needed": "yes before mixing with the primary baseline",
            "notes": "source_id=EasyDesign_2024; optional augmentation only in round2 baseline guidance.",
        },
        {
            "label_raw_name": "true value",
            "normalized_label": "diagnostic_activity_true_value_external_test",
            "path_type": "diagnostic_activity",
            "biological_meaning_cn": "Table S5 测试集中的实验真值/真实活性值。",
            "biological_meaning_en": "Experimental true value / true activity value in the Table S5 test dataset.",
            "assay_readout": "true activity value from CRISPR fluorescence reaction",
            "label_status": "measured",
            "trainable_as_primary_label": "yes_for_external_test_after_scale_review",
            "standard_unit": "Table S5 numeric scale; not confirmed identical to Table S3",
            "transform_needed": "yes before direct training/evaluation against Table S3 labels",
            "notes": "source_id=EasyDesign_2024; retain as external paper-test label, but scale compatibility remains unresolved.",
        },
    ]
    for raw in ["CNND", "CNN12a", "CNN12ae", "TransformerD", "Transformer12a", "Transformer12ae", "guide-expected-activities", "guide-activity", "mean-activity"]:
        rows.append(
            {
                "label_raw_name": raw,
                "normalized_label": "paper_or_tool_predicted_activity_score",
                "path_type": "predicted_library",
                "biological_meaning_cn": "模型或工具输出的预测活性分数；不是实验标签。",
                "biological_meaning_en": "Predicted activity score from a model or tool; not an experimental label.",
                "assay_readout": "prediction score",
                "label_status": "predicted",
                "trainable_as_primary_label": "no",
                "standard_unit": "model-specific prediction scale",
                "transform_needed": "not applicable as primary label",
                "notes": "source_id=EasyDesign_2024; keep separate from measured labels.",
            }
        )
    for raw in ["Normalized fluoresence", "Fluoresence value", "fluorescence unit", "activity in mismatch", "activity in different mutation scenarios"]:
        rows.append(
            {
                "label_raw_name": raw,
                "normalized_label": "figure_source_activity_metric",
                "path_type": "diagnostic_activity",
                "biological_meaning_cn": "图源数据中的荧光或活性相关指标；当前不是 tidy 训练标签。",
                "biological_meaning_en": "Fluorescence- or activity-related metric in figure source data; not a tidy training label at present.",
                "assay_readout": "figure source data metric",
                "label_status": "measured",
                "trainable_as_primary_label": "no",
                "standard_unit": "figure-specific scale",
                "transform_needed": "yes; requires manual mapping to crRNA-target rows",
                "notes": "source_id=EasyDesign_2024; inspect manually before any dataset use.",
            }
        )
    label = pd.concat([label, pd.DataFrame(rows)], ignore_index=True)
    write_xlsx(label, label_path, LABEL_COLUMNS, timestamp)
    outputs.append(safe_rel(label_path, paths.root))

    return outputs


def md_table(rows: list[dict[str, Any]], columns: list[str]) -> str:
    if not rows:
        return "_None._\n"
    lines = ["| " + " | ".join(columns) + " |", "| " + " | ".join(["---"] * len(columns)) + " |"]
    for row in rows:
        cells = []
        for col in columns:
            cells.append(clean_text(row.get(col, ""), 300).replace("|", "\\|"))
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def write_text(path: Path, title: str, lines: list[str], timestamp: str, backup_existing: bool = True) -> None:
    if backup_existing:
        backup(path, timestamp)
    path.write_text("# " + title + "\n\n" + "\n".join(lines).rstrip() + "\n", encoding="utf-8")


def append_or_replace_section(path: Path, marker: str, content: str, timestamp: str) -> None:
    backup(path, timestamp)
    start = f"<!-- BEGIN {marker} -->"
    end = f"<!-- END {marker} -->"
    section = f"{start}\n{content.rstrip()}\n{end}\n"
    old = path.read_text(encoding="utf-8", errors="replace") if path.exists() else ""
    pattern = re.compile(re.escape(start) + r".*?" + re.escape(end) + r"\n?", re.DOTALL)
    if pattern.search(old):
        new = pattern.sub(section, old)
    else:
        new = old.rstrip() + "\n\n" + section if old.strip() else section
    path.write_text(new, encoding="utf-8")


def write_markdown_reports(
    paths: Paths,
    primary_df: pd.DataFrame,
    augment_df: pd.DataFrame,
    feature_df: pd.DataFrame,
    stats: dict[str, Any],
    pdf_facts: list[dict[str, str]],
    raw_exports: dict[str, str],
    catalog_outputs: list[str],
    data_outputs: list[str],
    timestamp: str,
) -> list[str]:
    outputs: list[str] = []
    fact_rows_zh = [{"证据源": f["source"], "页码": f.get("page", ""), "结论": f["fact_zh"]} for f in pdf_facts]
    fact_rows_en = [{"source": f["source"], "page": f.get("page", ""), "fact": f["fact_en"]} for f in pdf_facts]
    image_rows_zh = [{"证据源": f["source"], "表格": f["fact_id"], "结论": f["fact_zh"]} for f in image_facts()]
    image_rows_en = [{"source": f["source"], "table": f["fact_id"], "fact": f["fact_en"]} for f in image_facts()]

    resolved_zh = [
        "`30 min`：PDF 方法部分确认 30 分钟荧光值被选为活性指标；本轮作为 Table S3 内部 baseline 的主标签。",
        "`20 min normalized`：PDF 确认它是 20 分钟读数归一到 30 分钟的派生/增强标签；不并入主标签。",
        "`out_logk_measurement`：对应 Table S4 增强数据；本轮单独保存为 optional augmentation，不进入默认 baseline。",
        "`true value`：对应 Table S5 测试集实验真值；保留为 external paper test，但与 Table S3 数值尺度仍需确认。",
        "standalone workbook 与 combined source-data workbook：以 combined workbook 为权威来源，standalone 表视为重复来源。",
        "figure source data：保留为证据或 metadata，不直接作为 tidy crRNA-target 训练行。",
        "`guide-expected-activities` 和模型列：全部归为 predicted score，不是 primary training label。",
        "PAM：PDF 支持 TTTN PAM；Excel 无独立 PAM 列，本轮只从序列 5' 端 TTTN 前缀推断并标记为 inferred。",
        "Fig.S3：属于特征/活性相关的聚合图源数据，不作为逐条训练标签。",
    ]
    resolved_en = [
        "`30 min`: the PDF methods confirm that the 30-minute fluorescence value was selected as the activity indicator; this round uses it as the primary label for the Table S3 internal baseline.",
        "`20 min normalized`: the PDF confirms that it is a derived/augmented label from the 20-minute readout normalized to 30 minutes; it is not merged into the primary label.",
        "`out_logk_measurement`: it belongs to the Table S4 augmented dataset; this round stores it separately as optional augmentation and excludes it from the default baseline.",
        "`true value`: it belongs to the Table S5 test dataset as the experimental true value; it is retained as an external paper test label, but its numeric scale versus Table S3 still needs confirmation.",
        "Standalone workbook versus combined source-data workbook: the combined workbook is used as the authoritative source, while standalone tables are treated as duplicate sources.",
        "Figure source data: retained as evidence or metadata, not used directly as tidy crRNA-target training rows.",
        "`guide-expected-activities` and model columns: all are classified as predicted scores, not primary training labels.",
        "PAM: the PDF supports a TTTN PAM; Excel has no separate PAM column, so this round only infers a 5-prime TTTN prefix and marks it as inferred.",
        "Fig.S3: it is aggregate figure source data about sequence/activity features and is not used as row-level training labels.",
    ]
    unresolved_zh = [
        "Table S3 的负值 `30 min` 标签与 Table S5 的正值 `true value` 是否可通过已知转换映射到同一尺度，仍需人工确认或方法学决定。",
        "Table S3 原始 `guide_target_hamming_dist` 与从 `guide_seq`/`target_at_guide` 直接计算的 mismatch 并不总一致；该原始列的真实语义仍需确认。",
        "Excel 中没有明确独立 PAM 列；当前 PAM 仅根据 TTTN 前缀推断，后续若要作为模型特征需要确认序列定义。",
        "Table S5 有少数 DNA context 不是 45 nt；这些行使用 best-match fallback 定位 target window，后续应人工复核。",
        "Table S4 增强数据是否进入首个正式训练流程，需要在模型方案确定后决定。",
        "Table S3 只有 type1/type2，不含具体 pathogen 名称；如果要做物种分组验证，需要额外映射。",
    ]
    unresolved_en = [
        "Whether the negative `30 min` labels in Table S3 and the positive `true value` labels in Table S5 can be mapped to one shared scale remains unresolved and needs either user confirmation or a methodological decision.",
        "The raw `guide_target_hamming_dist` in Table S3 does not always match the mismatch count computed directly from `guide_seq` and `target_at_guide`; the exact meaning of the raw column still needs confirmation.",
        "Excel has no explicit separate PAM column; the current PAM value is inferred only from the TTTN prefix and should be confirmed before use as a model feature.",
        "A small number of Table S5 DNA contexts are not 45 nt; those rows use a best-match fallback to locate the target window and should be manually reviewed later.",
        "Whether Table S4 augmentation should enter the first formal training workflow should be decided after the model plan is selected.",
        "Table S3 contains only type1/type2 and no exact pathogen names; species-level grouping validation needs an additional mapping.",
    ]

    outputs_generated = catalog_outputs + data_outputs

    audit_zh = [
        "## 范围",
        f"本轮只处理 `{SOURCE_ID}`。`01_raw` 未被修改；没有训练模型；没有合并不同 label system。",
        "",
        "## PDF 证据",
        md_table(fact_rows_zh, ["证据源", "页码", "结论"]),
        "## 截图/补充资料证据",
        md_table(image_rows_zh, ["证据源", "表格", "结论"]),
        "## 已解决问题",
        "\n".join(f"- {x}" for x in resolved_zh),
        "",
        "## 仍需解决的问题",
        "\n".join(f"- {x}" for x in unresolved_zh),
        "",
        "## 第二轮数据整理结论",
        f"- Table S3 作为主 baseline 内部分割来源：baseline_train={stats['baseline_train_rows']}，baseline_validation={stats['baseline_validation_rows']}。",
        f"- Table S5 保留为 external paper test：{stats['external_test_rows']} 行，但标记为 scale_unconfirmed。",
        f"- Table S4 作为 optional augmentation：{stats['augment_rows']} 行，不进入默认 baseline。",
    ]
    audit_en = [
        "## Scope",
        f"This round only processed `{SOURCE_ID}`. `01_raw` was not modified; no model was trained; different label systems were not merged.",
        "",
        "## PDF Evidence",
        md_table(fact_rows_en, ["source", "page", "fact"]),
        "## Screenshot / Supporting-Information Evidence",
        md_table(image_rows_en, ["source", "table", "fact"]),
        "## Resolved Questions",
        "\n".join(f"- {x}" for x in resolved_en),
        "",
        "## Remaining Questions",
        "\n".join(f"- {x}" for x in unresolved_en),
        "",
        "## Round-2 Data Organization Conclusions",
        f"- Table S3 is used for the primary internal baseline split: baseline_train={stats['baseline_train_rows']}, baseline_validation={stats['baseline_validation_rows']}.",
        f"- Table S5 is retained as the external paper test: {stats['external_test_rows']} rows, marked as scale_unconfirmed.",
        f"- Table S4 is optional augmentation: {stats['augment_rows']} rows, excluded from the default baseline.",
    ]
    audit_zh_path = paths.notes / "EasyDesign_2024_round2_data_audit_zh.md"
    audit_en_path = paths.notes / "EasyDesign_2024_round2_data_audit_en.md"
    write_text(audit_zh_path, "EasyDesign_2024 第二轮数据审计", audit_zh, timestamp)
    write_text(audit_en_path, "EasyDesign_2024 Round-2 Data Audit", audit_en, timestamp)
    outputs += [safe_rel(audit_zh_path, paths.root), safe_rel(audit_en_path, paths.root)]

    run_zh = [
        "## 范围",
        f"输入数据源：`01_raw/{SOURCE_ID}`、原论文 PDF、补充资料截图。",
        "",
        "## 本轮生成文件",
        "\n".join(f"- `{x}`" for x in outputs_generated),
        "",
        "## 分类判断",
        "- Table S3：diagnostic_activity，高优先级，内部 baseline 主表。",
        "- Table S4：diagnostic_activity，高优先级，但作为 optional augmentation。",
        "- Table S5：diagnostic_activity，高优先级，external paper test；预测列不作为标签。",
        "- Table S7/S8：experimental validation metadata。",
        "- Fig.S3：metadata/figure source data，不作为训练数据。",
        "",
        "## 数据质量检查",
        f"- primary v0 总行数：{len(primary_df)}。",
        f"- feature table 总行数：{len(feature_df)}。",
        f"- Table S5 target window 方法统计：{json.dumps(stats['table_s5_window_method_counts'], ensure_ascii=False)}。",
        f"- Table S5 DNA context 长度统计：{json.dumps(stats['table_s5_context_length_counts'], ensure_ascii=False)}。",
        f"- Table S5 target window 起点统计：{json.dumps(stats['table_s5_best_window_start_counts'], ensure_ascii=False)}。",
        f"- Table S3 中 unique target_at_guide：{stats['table_s3_unique_target_sequences']}。",
        "",
        "## 下一步建议",
        "- 第一次 baseline 先使用 `baseline_split` 中的 baseline_train/baseline_validation。",
        "- 不要把 Table S4 加入默认 baseline，除非显式启用 augmentation。",
        "- 在确认 label transform 前，不要把 Table S5 与 Table S3 作为同尺度评价。",
    ]
    run_en = [
        "## Scope",
        f"Input sources: `01_raw/{SOURCE_ID}`, the original paper PDF, and the supporting-information screenshot.",
        "",
        "## Files Generated In This Round",
        "\n".join(f"- `{x}`" for x in outputs_generated),
        "",
        "## Classification Decisions",
        "- Table S3: diagnostic_activity, high priority, primary internal-baseline table.",
        "- Table S4: diagnostic_activity, high priority, but optional augmentation.",
        "- Table S5: diagnostic_activity, high priority, external paper test; prediction columns are not labels.",
        "- Table S7/S8: experimental validation metadata.",
        "- Fig.S3: metadata/figure source data, not training data.",
        "",
        "## Data Quality Checks",
        f"- Primary v0 rows: {len(primary_df)}.",
        f"- Feature table rows: {len(feature_df)}.",
        f"- Table S5 target-window method counts: {json.dumps(stats['table_s5_window_method_counts'], ensure_ascii=False)}.",
        f"- Table S5 DNA context length counts: {json.dumps(stats['table_s5_context_length_counts'], ensure_ascii=False)}.",
        f"- Table S5 target-window start counts: {json.dumps(stats['table_s5_best_window_start_counts'], ensure_ascii=False)}.",
        f"- Unique Table S3 target_at_guide sequences: {stats['table_s3_unique_target_sequences']}.",
        "",
        "## Next Recommended Actions",
        "- Run the first baseline on `baseline_split` values baseline_train/baseline_validation.",
        "- Do not include Table S4 in the default baseline unless augmentation is explicitly enabled.",
        "- Do not treat Table S5 and Table S3 as one shared numeric label scale until a label transform is confirmed.",
    ]
    run_zh_path = paths.notes / "run_report_EasyDesign_2024_round2_zh.md"
    run_en_path = paths.notes / "run_report_EasyDesign_2024_round2_en.md"
    write_text(run_zh_path, "EasyDesign_2024 第二轮运行报告", run_zh, timestamp)
    write_text(run_en_path, "EasyDesign_2024 Round-2 Run Report", run_en, timestamp)
    outputs += [safe_rel(run_zh_path, paths.root), safe_rel(run_en_path, paths.root)]

    evidence_zh = [
        "## 证据追踪",
        "每条判断均区分证据和处理决定。",
        "",
        "### 决定 1：Table S3 是主 baseline 内部分割来源",
        "证据：PDF 说明 30 分钟荧光值作为活性指标；补充资料说明 Table S3 是 CRISPR fluorescence reaction 训练数据。",
        "处理：使用 `30 min`，按 target sequence hash 生成 baseline_train/baseline_validation。",
        "",
        "### 决定 2：Table S4 是 optional augmentation",
        "证据：PDF 说明 20/30 分钟读数用于数据增强；补充资料说明 Table S4 是 augment dataset。",
        "处理：单独输出 optional 文件，不并入默认 baseline。",
        "",
        "### 决定 3：Table S5 是 external paper test",
        "证据：补充资料说明 Table S5 是测试数据；PDF 方法部分给出 1,358 testing pairs。",
        "处理：保留 `true value` 和 paper prediction columns，但 prediction columns 只作参考。",
        "",
        "### 决定 4：PAM 只作推断字段",
        "证据：PDF 讨论 TTTN PAM；Excel 无独立 PAM 列。",
        "处理：从 5' TTTN 前缀推断 `pam`，并记录 `pam_inference_status`。",
    ]
    evidence_en = [
        "## Evidence Trace",
        "Each decision separates evidence from the processing action.",
        "",
        "### Decision 1: Table S3 is the primary internal-baseline source",
        "Evidence: the PDF states that the 30-minute fluorescence value was used as the activity indicator; the supplement describes Table S3 as CRISPR fluorescence-reaction training data.",
        "Action: use `30 min` and create baseline_train/baseline_validation by target-sequence hash.",
        "",
        "### Decision 2: Table S4 is optional augmentation",
        "Evidence: the PDF describes 20/30-minute readouts for data augmentation; the supplement describes Table S4 as the augment dataset.",
        "Action: output a separate optional file and exclude it from the default baseline.",
        "",
        "### Decision 3: Table S5 is the external paper test",
        "Evidence: the supplement describes Table S5 as test data; the PDF methods give 1,358 testing pairs.",
        "Action: keep `true value` and paper prediction columns, but use prediction columns only as references.",
        "",
        "### Decision 4: PAM is an inferred field only",
        "Evidence: the PDF discusses a TTTN PAM; Excel has no separate PAM column.",
        "Action: infer `pam` from the 5-prime TTTN prefix and record `pam_inference_status`.",
    ]
    evidence_zh_path = paths.notes / "evidence_trace_EasyDesign_2024_round2_zh.md"
    evidence_en_path = paths.notes / "evidence_trace_EasyDesign_2024_round2_en.md"
    write_text(evidence_zh_path, "EasyDesign_2024 第二轮证据追踪", evidence_zh, timestamp)
    write_text(evidence_en_path, "EasyDesign_2024 Round-2 Evidence Trace", evidence_en, timestamp)
    outputs += [safe_rel(evidence_zh_path, paths.root), safe_rel(evidence_en_path, paths.root)]

    method_zh = [
        "原始补充文件保持不变。第二轮审计使用原论文 PDF 和补充资料截图校对上一轮 unresolved questions。",
        "Excel 数据以 combined source-data workbook 为权威来源读取；standalone 表只作为重复来源记录。",
        "Table S3、Table S4、Table S5 分别被识别为训练、增强和测试数据。Table S3 的 `30 min` 被保留为主 baseline 内部标签，Table S4 的 `out_logk_measurement` 被保留为可选增强标签，Table S5 的 `true value` 被保留为外部测试真值。",
        "Table S5 的 DNA context 优先按论文描述的 10 nt flank 规则定位 25 nt target window；少数非 45 nt context 使用 best-match fallback。仅计算基础序列特征；未训练模型。",
    ]
    method_en = [
        "Raw supplementary files were left unchanged. The second-round audit used the original paper PDF and the supporting-information screenshot to resolve round-1 unresolved questions.",
        "Excel data were read from the combined source-data workbook as the authoritative source; standalone tables were recorded only as duplicate sources.",
        "Tables S3, S4, and S5 were identified as training, augmentation, and test data, respectively. The Table S3 `30 min` field was retained as the primary internal-baseline label, the Table S4 `out_logk_measurement` field was retained as an optional augmentation label, and the Table S5 `true value` field was retained as an external test truth value.",
        "Each Table S5 DNA context was mapped to a 25-nt target window by the paper-described 10-nt-flank rule when possible; non-45-nt contexts used a best-match fallback. Only basic sequence features were computed; no model was trained.",
    ]
    method_zh_path = paths.notes / "method_notes_EasyDesign_2024_round2_zh.md"
    method_en_path = paths.notes / "method_notes_EasyDesign_2024_round2_en.md"
    write_text(method_zh_path, "EasyDesign_2024 第二轮方法记录", method_zh, timestamp)
    write_text(method_en_path, "EasyDesign_2024 Round-2 Method Notes", method_en, timestamp)
    outputs += [safe_rel(method_zh_path, paths.root), safe_rel(method_en_path, paths.root)]

    problems_zh = [
        "## 已解决",
        "\n".join(f"- {x}" for x in resolved_zh),
        "",
        "## 仍需确认",
        "\n".join(f"- {x}" for x in unresolved_zh),
    ]
    problems_en = [
        "## Resolved",
        "\n".join(f"- {x}" for x in resolved_en),
        "",
        "## Still Needs Confirmation",
        "\n".join(f"- {x}" for x in unresolved_en),
    ]
    problems_zh_path = paths.notes / "problems_to_resolve_zh.md"
    problems_en_path = paths.notes / "problems_to_resolve_en.md"
    write_text(problems_zh_path, "待解决问题 EasyDesign_2024 第二轮", problems_zh, timestamp)
    write_text(problems_en_path, "Problems To Resolve EasyDesign_2024 Round 2", problems_en, timestamp)
    outputs += [safe_rel(problems_zh_path, paths.root), safe_rel(problems_en_path, paths.root)]

    legacy_section = "## EasyDesign_2024 Round2\n\n" + "\n".join(f"- {x}" for x in unresolved_zh)
    append_or_replace_section(paths.notes / "problems_to_resolve.md", "EasyDesign_2024_ROUND2", legacy_section, timestamp)
    outputs.append("99_notes/problems_to_resolve.md")

    usage_zh = [
        "## 文件位置选择",
        "这份文档放在 `04_candidate_ml_dataset/`，因为它直接说明候选建模数据如何进入第一次 baseline workflow。原始提取表仍在 `02_extracted_tables/`，最小清洗表在 `03_cleaned_minimal/`，但模型训练入口应从 `04_candidate_ml_dataset/` 读取。",
        "",
        "## 推荐的第一次 baseline 输入",
        "- 默认训练/验证文件：`diagnostic_activity_feature_table_v0.csv`。",
        "- 只使用 `label_is_primary_baseline == yes` 的行。",
        "- 训练集：`baseline_split == baseline_train`。",
        "- 验证集：`baseline_split == baseline_validation`。",
        "- 标签列：`label_normalized`，但只在 `label_scale_group == table_s3_log_or_transformed_30min_activity` 内使用。",
        "",
        "## 暂不默认使用的数据",
        "- `baseline_split == external_test_scale_unconfirmed`：这是 Table S5 外部测试集，保留 `true value`，但数值尺度与 Table S3 尚未确认。",
        "- `diagnostic_activity_augmented_optional_v0.csv`：这是 Table S4 增强数据，只有显式启用 augmentation 时才使用。",
        "- 所有 `paper_prediction_*` 列：这是论文模型预测值，只能用于复现实验对照或 sanity check，不是 label。",
        "",
        "## 最低可运行 workflow",
        "1. 读取 `diagnostic_activity_feature_table_v0.csv`。",
        "2. 过滤 `label_is_primary_baseline == yes`。",
        "3. 使用 `crRNA_sequence`、`target_sequence` 和基础数值特征做 baseline 特征。",
        "4. 使用 `baseline_train` 训练，`baseline_validation` 验证。",
        "5. 暂不报告 Table S5 外部测试性能，除非先解决 label scale transform。",
    ]
    usage_en = [
        "## Why This File Lives Here",
        "This document is stored in `04_candidate_ml_dataset/` because it explains how candidate modeling data should enter the first baseline workflow. Raw extracted tables remain in `02_extracted_tables/`, and minimally cleaned tables remain in `03_cleaned_minimal/`, but model training should start from `04_candidate_ml_dataset/`.",
        "",
        "## Recommended First Baseline Inputs",
        "- Default training/validation file: `diagnostic_activity_feature_table_v0.csv`.",
        "- Use only rows where `label_is_primary_baseline == yes`.",
        "- Training set: `baseline_split == baseline_train`.",
        "- Validation set: `baseline_split == baseline_validation`.",
        "- Label column: `label_normalized`, but only within `label_scale_group == table_s3_log_or_transformed_30min_activity`.",
        "",
        "## Data Not Used By Default Yet",
        "- `baseline_split == external_test_scale_unconfirmed`: this is the Table S5 external test set with `true value`, but its numeric scale versus Table S3 is not confirmed.",
        "- `diagnostic_activity_augmented_optional_v0.csv`: this is the Table S4 augmented dataset and should only be used when augmentation is explicitly enabled.",
        "- All `paper_prediction_*` columns: these are paper model predictions and may be used only for reproduction checks or sanity checks, not as labels.",
        "",
        "## Minimal Runnable Workflow",
        "1. Read `diagnostic_activity_feature_table_v0.csv`.",
        "2. Filter `label_is_primary_baseline == yes`.",
        "3. Use `crRNA_sequence`, `target_sequence`, and the basic numeric features as baseline features.",
        "4. Train on `baseline_train` and validate on `baseline_validation`.",
        "5. Do not report Table S5 external test performance until the label-scale transform is resolved.",
    ]
    usage_zh_path = paths.candidate / "baseline_data_usage_guide_zh.md"
    usage_en_path = paths.candidate / "baseline_data_usage_guide_en.md"
    write_text(usage_zh_path, "Baseline 数据使用指南", usage_zh, timestamp)
    write_text(usage_en_path, "Baseline Data Usage Guide", usage_en, timestamp)
    outputs += [safe_rel(usage_zh_path, paths.root), safe_rel(usage_en_path, paths.root)]

    split_zh = [
        "## Split Plan",
        f"- Table S3 被划分为 baseline_train={stats['baseline_train_rows']} 行和 baseline_validation={stats['baseline_validation_rows']} 行。",
        "- 划分方法：对 `target_sequence` 做 SHA1 hash，并按 bucket < 80 进入训练集，其余进入验证集。",
        "- 目的：降低同一 target sequence 同时出现在训练和验证中的风险。",
        f"- Table S5 保留为 external_test_scale_unconfirmed={stats['external_test_rows']} 行；在 label scale 未确认前不作为默认测试集。",
        f"- Table S4 保留为 optional_augmentation={stats['augment_rows']} 行；默认 baseline 不使用。",
    ]
    split_en = [
        "## Split Plan",
        f"- Table S3 was split into baseline_train={stats['baseline_train_rows']} rows and baseline_validation={stats['baseline_validation_rows']} rows.",
        "- Method: compute a SHA1 hash bucket from `target_sequence`; bucket < 80 goes to training, and the rest goes to validation.",
        "- Purpose: reduce the risk that the same target sequence appears in both training and validation.",
        f"- Table S5 is retained as external_test_scale_unconfirmed={stats['external_test_rows']} rows; it is not the default test set until label scale is confirmed.",
        f"- Table S4 is retained as optional_augmentation={stats['augment_rows']} rows; the default baseline does not use it.",
    ]
    split_zh_path = paths.candidate / "split_plan_zh.md"
    split_en_path = paths.candidate / "split_plan_en.md"
    write_text(split_zh_path, "数据划分计划", split_zh, timestamp)
    write_text(split_en_path, "Split Plan", split_en, timestamp)
    write_text(paths.candidate / "split_plan.md", "数据划分计划", split_zh, timestamp)
    outputs += [safe_rel(split_zh_path, paths.root), safe_rel(split_en_path, paths.root), "04_candidate_ml_dataset/split_plan.md"]

    build_zh = [
        "## 输入",
        f"- 权威 workbook：`{safe_rel(paths.raw_combined, paths.root)}`。",
        "- 证据源：原论文 PDF、补充资料截图、上一轮 catalog。",
        "",
        "## 输出数据",
        f"- `diagnostic_activity_v0.csv`：{len(primary_df)} 行，包含 Table S3 primary internal baseline 与 Table S5 external paper test。",
        f"- `diagnostic_activity_feature_table_v0.csv`：{len(feature_df)} 行，包含基础序列特征。",
        f"- `diagnostic_activity_augmented_optional_v0.csv`：{len(augment_df)} 行，包含 Table S4 optional augmentation。",
        "",
        "## 保留/排除规则",
        "- 保留原始 record id、source_table_id、source_sheet、label_raw_name、label_scale_group。",
        "- 不把 30 min、20 min normalized、out_logk_measurement、true value 强行合并为同一标签。",
        "- 不把 paper model prediction columns 当作 label。",
    ]
    build_en = [
        "## Inputs",
        f"- Authoritative workbook: `{safe_rel(paths.raw_combined, paths.root)}`.",
        "- Evidence sources: original paper PDF, supporting-information screenshot, and round-1 catalog.",
        "",
        "## Output Data",
        f"- `diagnostic_activity_v0.csv`: {len(primary_df)} rows, containing the Table S3 primary internal baseline and the Table S5 external paper test.",
        f"- `diagnostic_activity_feature_table_v0.csv`: {len(feature_df)} rows, containing basic sequence features.",
        f"- `diagnostic_activity_augmented_optional_v0.csv`: {len(augment_df)} rows, containing the Table S4 optional augmentation.",
        "",
        "## Retention / Exclusion Rules",
        "- Keep original record id, source_table_id, source_sheet, label_raw_name, and label_scale_group.",
        "- Do not force 30 min, 20 min normalized, out_logk_measurement, and true value into one label.",
        "- Do not treat paper model prediction columns as labels.",
    ]
    build_zh_path = paths.candidate / "dataset_build_report_zh.md"
    build_en_path = paths.candidate / "dataset_build_report_en.md"
    write_text(build_zh_path, "数据集构建报告", build_zh, timestamp)
    write_text(build_en_path, "Dataset Build Report", build_en, timestamp)
    outputs += [safe_rel(build_zh_path, paths.root), safe_rel(build_en_path, paths.root)]

    return outputs


def main() -> None:
    root = project_root()
    paths = ensure_paths(root)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    pdf_facts = extract_pdf_facts()
    primary_df, augment_df, stats = build_primary_dataset(paths)
    feature_df = build_feature_table(primary_df)

    raw_exports = export_authoritative_raw_csvs(paths, timestamp)

    data_outputs = []
    output_map = {
        paths.cleaned / "diagnostic_activity_minimal.csv": primary_df,
        paths.cleaned / "diagnostic_activity_augmented_optional.csv": augment_df,
        paths.candidate / "diagnostic_activity_v0.csv": primary_df,
        paths.candidate / "diagnostic_activity_feature_table_v0.csv": feature_df,
        paths.candidate / "diagnostic_activity_augmented_optional_v0.csv": augment_df,
    }
    for out_path, df in output_map.items():
        write_csv_with_backup(df, out_path, timestamp)
        data_outputs.append(safe_rel(out_path, root))

    catalog_outputs = update_catalogs(paths, primary_df, augment_df, raw_exports, timestamp)

    report_outputs = write_markdown_reports(
        paths=paths,
        primary_df=primary_df,
        augment_df=augment_df,
        feature_df=feature_df,
        stats=stats,
        pdf_facts=pdf_facts,
        raw_exports=raw_exports,
        catalog_outputs=catalog_outputs,
        data_outputs=data_outputs,
        timestamp=timestamp,
    )

    result = {
        "source_id": SOURCE_ID,
        "stats": stats,
        "catalog_outputs": catalog_outputs,
        "data_outputs": data_outputs,
        "report_outputs": report_outputs,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
