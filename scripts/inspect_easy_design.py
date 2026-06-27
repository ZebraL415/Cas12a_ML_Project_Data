#!/usr/bin/env python3
"""Audit EasyDesign_2024 raw files without modifying 01_raw.

This script follows the project rule: audit first, extract raw tables second,
and do not build or train a model. It scans source documentation and every
worksheet in every .xlsx workbook under 01_raw/EasyDesign_2024, updates the
central catalog workbooks, exports only high-priority diagnostic activity
worksheets to raw CSV, and writes audit notes.
"""

from __future__ import annotations

import json
import math
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_ID = "EasyDesign_2024"
PAPER_SHORT = "EasyDesign"
YEAR = "2024"
FULL_TITLE = (
    "Enhanced crRNA design system with Deep learning for Cas12a-mediated Diagnostics"
)

CATALOG_COLUMNS = [
    "source_id",
    "paper_short",
    "year",
    "full_title",
    "path_type",
    "source_type",
    "raw_path",
    "file_names",
    "data_access_status",
    "record_unit",
    "label_type",
    "label_status",
    "sample_size_estimated",
    "sequence_fields_detected",
    "usable_for_training",
    "usable_for_extension",
    "priority",
    "main_risk",
    "notes",
]

SHEET_INDEX_COLUMNS = [
    "source_id",
    "file_name",
    "file_path",
    "sheet_name",
    "n_rows",
    "n_cols",
    "first_columns",
    "example_values",
    "guessed_content",
    "record_unit_guess",
    "has_crRNA_sequence",
    "has_target_sequence",
    "has_PAM",
    "has_label",
    "label_raw_candidates",
    "label_status_guess",
    "path_type_guess",
    "priority",
    "recommended_action",
    "extracted_csv_path",
    "notes",
    "evidence_source",
    "title_or_caption",
]

LABEL_COLUMNS = [
    "label_raw_name",
    "normalized_label",
    "path_type",
    "biological_meaning_cn",
    "biological_meaning_en",
    "assay_readout",
    "label_status",
    "trainable_as_primary_label",
    "standard_unit",
    "transform_needed",
    "notes",
]


@dataclass
class SheetAudit:
    source_id: str
    file_name: str
    file_path: str
    sheet_name: str
    n_rows: int
    n_cols: int
    header_row: int
    first_columns: list[str]
    example_values: list[dict[str, str]]
    title_or_caption: str
    guessed_content: str
    record_unit_guess: str
    has_crrna_sequence: str
    has_target_sequence: str
    has_pam: str
    has_label: str
    label_raw_candidates: list[str] = field(default_factory=list)
    label_status_guess: str = "unclear"
    path_type_guess: str = "unclear"
    priority: str = "low"
    recommended_action: str = "inspect_manually"
    extracted_csv_path: str = ""
    notes: str = ""
    evidence_source: str = ""

    def as_row(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "file_name": self.file_name,
            "file_path": self.file_path,
            "sheet_name": self.sheet_name,
            "n_rows": self.n_rows,
            "n_cols": self.n_cols,
            "first_columns": json.dumps(self.first_columns[:20], ensure_ascii=False),
            "example_values": json.dumps(self.example_values[:3], ensure_ascii=False),
            "guessed_content": self.guessed_content,
            "record_unit_guess": self.record_unit_guess,
            "has_crRNA_sequence": self.has_crrna_sequence,
            "has_target_sequence": self.has_target_sequence,
            "has_PAM": self.has_pam,
            "has_label": self.has_label,
            "label_raw_candidates": "; ".join(self.label_raw_candidates),
            "label_status_guess": self.label_status_guess,
            "path_type_guess": self.path_type_guess,
            "priority": self.priority,
            "recommended_action": self.recommended_action,
            "extracted_csv_path": self.extracted_csv_path,
            "notes": self.notes,
            "evidence_source": self.evidence_source,
            "title_or_caption": self.title_or_caption,
        }


def project_root_from_script() -> Path:
    return Path(__file__).resolve().parents[1]


def clean_cell(value: Any, max_len: int = 500) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def safe_rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def human_size(num_bytes: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    size = float(num_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{num_bytes} B"


def sanitize_id(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", text.strip())
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "sheet"


def ensure_dirs(root: Path) -> dict[str, Path]:
    paths = {
        "raw": root / "01_raw" / SOURCE_ID,
        "catalog": root / "00_data_catalog",
        "diagnostic": root / "02_extracted_tables" / "diagnostic_activity",
        "notes": root / "99_notes",
        "scripts": root / "scripts",
    }
    if not paths["raw"].exists():
        raise FileNotFoundError(f"Missing raw directory: {paths['raw']}")
    for key in ["catalog", "diagnostic", "notes", "scripts"]:
        paths[key].mkdir(parents=True, exist_ok=True)
    return paths


def backup_if_exists(path: Path, timestamp: str) -> Path | None:
    if not path.exists():
        return None
    backup = path.with_name(f"{path.stem}_{timestamp}_backup{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def possible_usage(path: Path) -> str:
    name = path.name.lower()
    suffix = path.suffix.lower()
    rel = str(path).lower()
    if name == ".ds_store":
        return "macOS metadata; not biologically informative."
    if suffix == ".xlsx":
        return "Excel supplementary/source data workbook for sheet-level audit."
    if name == "readme.md":
        return "Primary source documentation for tool purpose, inputs, and outputs."
    if name in {"requirements.txt", "environment.yml", "setup.py"}:
        return "Software dependency/package metadata; useful for reproducibility context."
    if suffix == ".py":
        return "Source code; useful for output field names and model/input semantics."
    if suffix == ".pb" or "models/" in rel:
        return "Saved model artifact; evidence of prediction models, not raw labels."
    if suffix in {".fasta", ".fa"}:
        return "Example sequence input."
    if name == "license":
        return "Repository license."
    return "Repository file; usage unclear from extension."


def file_kind(path: Path) -> str:
    suffix = path.suffix.lower()
    if path.name == ".DS_Store":
        return "macOS metadata"
    if suffix == ".xlsx":
        return "Excel workbook"
    if suffix == ".md":
        return "Markdown documentation"
    if suffix in {".txt", ".yml", ".yaml"}:
        return "Text/config"
    if suffix == ".py":
        return "Python source"
    if suffix == ".pb":
        return "TensorFlow SavedModel"
    if suffix in {".fasta", ".fa"}:
        return "FASTA"
    if path.name == "LICENSE":
        return "License"
    return suffix.lstrip(".") or "file"


def inventory_files(raw_dir: Path, root: Path) -> list[dict[str, str]]:
    rows = []
    for path in sorted(raw_dir.rglob("*")):
        if path.is_file():
            rows.append(
                {
                    "file_path": safe_rel(path, root),
                    "file_type": file_kind(path),
                    "file_size": human_size(path.stat().st_size),
                    "possible_usage": possible_usage(path),
                }
            )
    return rows


def read_text_file(path: Path, max_chars: int = 20000) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")[:max_chars]
    except OSError:
        return ""


def extract_doc_facts(raw_dir: Path, root: Path) -> tuple[list[dict[str, str]], list[str]]:
    doc_paths = sorted(
        [
            p
            for p in raw_dir.rglob("*")
            if p.is_file()
            and (
                p.name.lower().startswith("readme")
                or p.suffix.lower() in {".md", ".txt", ".yml", ".yaml"}
                or "supp" in p.name.lower()
                or "description" in p.name.lower()
            )
        ]
    )
    keyword_re = re.compile(
        r"activity|fluorescence|rfu|guide|crrna|target|sequence|model|train|test|"
        r"predict|score|diagnostic|cas12a|rpa|output|input",
        re.IGNORECASE,
    )
    facts: list[dict[str, str]] = []
    for path in doc_paths:
        text = read_text_file(path)
        for line_no, line in enumerate(text.splitlines(), 1):
            cleaned = clean_cell(line, max_len=900)
            if cleaned and keyword_re.search(cleaned):
                facts.append(
                    {
                        "file_path": safe_rel(path, root),
                        "line": str(line_no),
                        "evidence": cleaned,
                    }
                )
    return facts, [safe_rel(p, root) for p in doc_paths]


def row_values(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, max_cols: int) -> list[str]:
    return [
        clean_cell(ws.cell(row=row_idx, column=col).value)
        for col in range(1, min(max_cols, ws.max_column or 1) + 1)
    ]


def nonempty_count(values: list[str]) -> int:
    return sum(1 for value in values if value)


def header_score(values: list[str]) -> float:
    """Score a top-row candidate as a plausible table header.

    Supplementary sheets often contain one title row, one panel-letter row,
    then a true header. Choosing the densest row alone can accidentally pick
    the first data row, so this favors semantic header words and penalizes
    numeric/value-like rows.
    """
    keywords = re.compile(
        r"no\.?$|name|pathogen|species|type|sequence|guide|crrna|target|dna|"
        r"value|measurement|activity|fluoresc|fluoresence|rfu|min|model|"
        r"position|primer|template|templat|mismatch|coefficient|control|"
        r"sample|fold|predicted|true|total|number",
        re.IGNORECASE,
    )
    nonempty = [value for value in values if value]
    if not nonempty:
        return -1
    panel_like = sum(1 for value in nonempty if re.fullmatch(r"\(?[A-Z]\)?", value))
    numeric_like = sum(1 for value in nonempty if re.fullmatch(r"[-+]?\d+(\.\d+)?", value))
    keyword_hits = sum(1 for value in nonempty if keywords.search(value))
    alpha_hits = sum(1 for value in nonempty if re.search(r"[A-Za-z]", value))
    sequence_like = sum(1 for value in nonempty if re.fullmatch(r"[ACGTUacgtu\s-]{12,}", value))
    return (keyword_hits * 6) + alpha_hits + len(nonempty) - (numeric_like * 3) - (panel_like * 5) - (sequence_like * 2)


def find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet, top_scan_rows: int = 12
) -> tuple[int, str]:
    max_cols = min(ws.max_column or 1, 80)
    candidates: list[tuple[float, int, int, list[str]]] = []
    title = ""
    for row_idx in range(1, min(ws.max_row or 1, top_scan_rows) + 1):
        values = row_values(ws, row_idx, max_cols)
        count = nonempty_count(values)
        if not title:
            first_nonempty = next((v for v in values if v), "")
            if first_nonempty:
                title = first_nonempty
        if count == 0:
            continue
        first_nonempty = next((v for v in values if v), "")
        looks_like_caption = bool(
            re.match(r"^(table|fig|source data)", first_nonempty, re.IGNORECASE)
            and count <= max(2, max_cols // 8)
        )
        if looks_like_caption:
            continue
        score = header_score(values)
        candidates.append((score, count, row_idx, values))
    if not candidates:
        return 1, title
    candidates.sort(key=lambda item: (-item[0], -item[1], item[2]))
    return candidates[0][2], title


def headers_from_row(values: list[str]) -> list[str]:
    headers = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values, 1):
        header = clean_cell(value, max_len=120)
        if not header:
            header = f"blank_{idx}"
        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)
    return headers


def examples_after_header(
    ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, headers: list[str]
) -> list[dict[str, str]]:
    examples: list[dict[str, str]] = []
    max_cols = min(len(headers), 20)
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=min(ws.max_row or header_row, header_row + 8),
        max_col=max_cols,
        values_only=True,
    ):
        cleaned = [clean_cell(v, max_len=180) for v in row]
        if not any(cleaned):
            continue
        examples.append({headers[i]: cleaned[i] for i in range(len(cleaned))})
        if len(examples) >= 3:
            break
    return examples


def all_detection_text(
    sheet_name: str,
    title: str,
    headers: list[str],
    examples: list[dict[str, str]],
) -> str:
    return " ".join(
        [
            sheet_name,
            title,
            " ".join(headers),
            json.dumps(examples, ensure_ascii=False),
        ]
    ).lower()


def find_label_candidates(headers: list[str]) -> list[str]:
    candidates: list[str] = []
    label_patterns = [
        r"\b30\s*min\b",
        r"\b20\s*min\b",
        r"true value",
        r"out_logk",
        r"logk",
        r"activity",
        r"fluoresc",
        r"fluoresence",
        r"\brfu\b",
        r"expected",
        r"predicted",
        r"prediction",
        r"\bcnnd\b",
        r"\bcnn12a\b",
        r"\bcnn12ae\b",
        r"transformer",
        r"coefficient",
        r"spearman",
        r"pearson",
    ]
    for header in headers:
        lower = header.lower()
        if any(re.search(pattern, lower) for pattern in label_patterns):
            # Avoid counting row-number or metadata fields as labels.
            if lower.strip() in {"no.", "no", "number", "template no."}:
                continue
            candidates.append(header)
    return list(dict.fromkeys(candidates))


def classify_sheet(
    file_name: str,
    sheet_name: str,
    title: str,
    headers: list[str],
    examples: list[dict[str, str]],
) -> dict[str, Any]:
    text = all_detection_text(sheet_name, title, headers, examples)
    sequence_detection_text = " ".join(
        [sheet_name, " ".join(headers), json.dumps(examples, ensure_ascii=False)]
    ).lower()
    header_text = " ".join(headers).lower()

    has_crrna = (
        "yes"
        if (
            re.search(r"\bguide_seq\b|guide[-_\s]?target[-_\s]?sequences", sequence_detection_text)
            or re.search(r"\bcrrna\b", header_text)
            or ("crrna" in sheet_name.lower() and "sequence" in header_text)
            or ("guide" in header_text and "sequence" in header_text)
        )
        else "no"
    )
    if "no. of crrna" in header_text:
        has_crrna = "no"
    if has_crrna == "no" and "guide" in sequence_detection_text and "sequence" in sequence_detection_text:
        has_crrna = "unclear"

    has_target = (
        "yes"
        if re.search(
            r"target_at_guide|target[-_\s]?sequence|target[-_\s]?at|template|templat|dna\b|guide-target-sequences",
            sequence_detection_text,
        )
        else "no"
    )
    if has_target == "no" and "sequence" in sequence_detection_text and re.search(r"pathogen|template|templat|dna", sequence_detection_text):
        has_target = "unclear"

    has_pam = "yes" if re.search(r"\bpam\b", sequence_detection_text) else "no"
    label_candidates = find_label_candidates(headers)
    has_label = "yes" if label_candidates else "no"

    measured_terms = [
        "30 min",
        "20 min normalized",
        "true value",
        "out_logk",
        "fluorescence",
        "fluoresence",
        "rfu",
        "normalized fluores",
        "fluorescence unit",
    ]
    predicted_terms = [
        "expected",
        "predicted",
        "prediction",
        "cnnd",
        "cnn12a",
        "cnn12ae",
        "transformer",
        "guide-expected-activities",
        "model",
    ]
    performance_terms = ["spearman", "pearson", "cross-validation", "coefficient", "fold"]
    primer_terms = ["primer", "rpa primer"]
    species_terms = ["pathogen", "species", "type", "no. of crrna"]
    figure_terms = ["fig.", "source data figure", "figure"]
    specificity_re = re.compile(r"\b(specificity|mutant|wt|wild type|ratio|discrimination)\b", re.IGNORECASE)

    measured_hit = any(term in text for term in measured_terms)
    predicted_hit = any(term in text for term in predicted_terms)
    performance_hit = any(term in text for term in performance_terms)
    primer_hit = any(term in text for term in primer_terms)
    species_hit = any(term in text for term in species_terms)
    figure_hit = any(term in text for term in figure_terms)
    specificity_hit = bool(specificity_re.search(text))

    if has_label == "yes" and measured_hit:
        label_status = "measured"
    elif has_label == "yes" and predicted_hit:
        label_status = "predicted"
    elif has_label == "yes":
        label_status = "unclear"
    elif "annotation" in text:
        label_status = "annotation"
    else:
        label_status = "metadata"

    record_unit = "unclear"
    path_type = "unclear"
    guessed_content = "unclear sheet content"
    priority = "low"
    action = "inspect_manually"
    notes: list[str] = []

    if performance_hit:
        path_type = "metadata_only"
        record_unit = "metadata"
        guessed_content = "model performance or figure summary metadata"
        priority = "low"
        action = "keep_as_metadata"
    elif ("template no" in header_text or "nucleic acid templates" in text) and "sequence" in header_text:
        path_type = "metadata_only"
        record_unit = "genomic target site"
        guessed_content = "pathogen/template sequence metadata"
        priority = "medium"
        action = "keep_as_metadata"
    elif specificity_hit and measured_hit:
        path_type = "snv_specificity"
        record_unit = "SNV variant"
        guessed_content = "specificity or mismatch-related measured activity"
        priority = "medium"
        action = "inspect_manually"
    elif has_crrna == "yes" and has_target == "yes" and measured_hit:
        path_type = "diagnostic_activity"
        record_unit = "crRNA-target pair"
        guessed_content = "measured diagnostic activity table with guide/target sequences"
        priority = "high"
        action = "extract_to_csv"
    elif measured_hit and figure_hit:
        path_type = "diagnostic_activity"
        record_unit = "figure source data"
        guessed_content = "figure source data containing fluorescence/activity-like measurements"
        priority = "medium"
        action = "inspect_manually"
        notes.append("Contains measurement-like fields but lacks a tidy crRNA-target pair structure.")
    elif predicted_hit and has_crrna in {"yes", "unclear"}:
        path_type = "predicted_library"
        record_unit = "crRNA-target pair" if has_target == "yes" else "unclear"
        guessed_content = "predicted activity or model output table"
        priority = "medium"
        action = "inspect_manually"
    elif primer_hit and has_crrna in {"yes", "unclear"}:
        path_type = "predicted_library"
        record_unit = "primer"
        guessed_content = "designed crRNA and/or RPA primer table"
        priority = "medium"
        action = "keep_as_metadata"
    elif "crrna" in text and "sequence" in text:
        path_type = "predicted_library"
        record_unit = "crRNA-target pair" if has_target == "yes" else "genomic target site"
        guessed_content = "candidate crRNA sequence library without measured label"
        priority = "medium"
        action = "keep_as_metadata"
    elif "template" in text or "templat" in text or ("sequence" in header_text and species_hit):
        path_type = "metadata_only"
        record_unit = "genomic target site"
        guessed_content = "pathogen/template sequence metadata"
        priority = "medium"
        action = "keep_as_metadata"
    elif species_hit:
        path_type = "metadata_only"
        record_unit = "metadata"
        guessed_content = "pathogen/species metadata"
        priority = "medium"
        action = "keep_as_metadata"

    if label_status == "predicted" and priority == "high":
        priority = "medium"
        action = "inspect_manually"
        notes.append("Predicted score detected; do not treat as primary measured label.")
    if predicted_hit and measured_hit:
        notes.append("Sheet includes measured-like and prediction/model-score fields; keep them separate.")
    if "imt2214-sup-0002" in file_name.lower() or file_name.lower().startswith("table s"):
        notes.append("May duplicate a standalone table or a combined source-data workbook; do not merge automatically.")

    evidence_parts = []
    if title:
        evidence_parts.append(f"title/caption: {title}")
    evidence_parts.append(f"sheet_name: {sheet_name}")
    evidence_parts.append("columns: " + ", ".join(headers[:20]))

    return {
        "guessed_content": guessed_content,
        "record_unit_guess": record_unit,
        "has_crrna_sequence": has_crrna,
        "has_target_sequence": has_target,
        "has_pam": has_pam,
        "has_label": has_label,
        "label_raw_candidates": label_candidates,
        "label_status_guess": label_status,
        "path_type_guess": path_type,
        "priority": priority,
        "recommended_action": action,
        "notes": " ".join(notes),
        "evidence_source": " | ".join(evidence_parts),
    }


def scan_workbooks(raw_dir: Path, root: Path) -> list[SheetAudit]:
    audits: list[SheetAudit] = []
    for workbook_path in sorted(raw_dir.rglob("*.xlsx")):
        wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
        try:
            for ws in wb.worksheets:
                n_rows = ws.max_row or 0
                n_cols = ws.max_column or 0
                header_row, title = find_header_row(ws)
                header_values = row_values(ws, header_row, min(n_cols or 1, 80))
                headers = headers_from_row(header_values)
                examples = examples_after_header(ws, header_row, headers)
                classification = classify_sheet(workbook_path.name, ws.title, title, headers, examples)
                audit = SheetAudit(
                    source_id=SOURCE_ID,
                    file_name=workbook_path.name,
                    file_path=safe_rel(workbook_path, root),
                    sheet_name=ws.title,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    header_row=header_row,
                    first_columns=[
                        f"{get_column_letter(i + 1)}: {name}" for i, name in enumerate(headers[:20])
                    ],
                    example_values=examples,
                    title_or_caption=title,
                    **classification,
                )
                audits.append(audit)
        finally:
            wb.close()
    return audits


def read_existing_xlsx(path: Path, columns: list[str]) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=columns)
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception:
        return pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns].fillna("")


def style_xlsx(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 10
            for cell in list(col)[:200]:
                value = clean_cell(cell.value, max_len=200)
                max_len = max(max_len, min(len(value), 60))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)
        ws.row_dimensions[1].height = 28
    wb.save(path)


def write_xlsx(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    out = out[columns].fillna("")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=path.stem[:31])
    style_xlsx(path)


def update_sheet_index(audits: list[SheetAudit], catalog_dir: Path, timestamp: str) -> Path:
    path = catalog_dir / "source_sheet_index.xlsx"
    backup_if_exists(path, timestamp)
    existing = read_existing_xlsx(path, SHEET_INDEX_COLUMNS)
    existing = existing[existing["source_id"] != SOURCE_ID]
    new_df = pd.DataFrame([audit.as_row() for audit in audits], columns=SHEET_INDEX_COLUMNS)
    combined = pd.concat([existing, new_df], ignore_index=True)
    write_xlsx(combined, path, SHEET_INDEX_COLUMNS)
    return path


def high_priority_sample_size(audits: list[SheetAudit]) -> str:
    parts = []
    for audit in audits:
        if audit.priority == "high" and audit.path_type_guess == "diagnostic_activity":
            rows = max(audit.n_rows - audit.header_row, 0)
            parts.append(f"{audit.file_name}:{audit.sheet_name}≈{rows}")
    return "; ".join(parts)


def update_master_catalog(
    audits: list[SheetAudit], inventory: list[dict[str, str]], catalog_dir: Path, timestamp: str
) -> Path:
    path = catalog_dir / "master_data_catalog.xlsx"
    backup_if_exists(path, timestamp)
    existing = read_existing_xlsx(path, CATALOG_COLUMNS)
    existing = existing[existing["source_id"] != SOURCE_ID]
    xlsx_files = [row["file_path"] for row in inventory if row["file_type"] == "Excel workbook"]
    detected_sequence_fields = sorted(
        {
            candidate
            for audit in audits
            for candidate in json.loads(audit.as_row()["first_columns"])
            if re.search(r"guide|crrna|target|dna|sequence|template", candidate, re.IGNORECASE)
        }
    )
    row = {
        "source_id": SOURCE_ID,
        "paper_short": PAPER_SHORT,
        "year": YEAR,
        "full_title": FULL_TITLE,
        "path_type": "diagnostic_activity",
        "source_type": "GitHub repository plus supplementary Excel/source-data workbooks and saved prediction models",
        "raw_path": f"01_raw/{SOURCE_ID}",
        "file_names": "; ".join(xlsx_files),
        "data_access_status": "local raw files available; 01_raw not modified",
        "record_unit": "crRNA-target pair; pathogen/template metadata; figure source data",
        "label_type": "CRISPR fluorescence/logK diagnostic activity; auxiliary predicted model scores",
        "label_status": "measured",
        "sample_size_estimated": high_priority_sample_size(audits),
        "sequence_fields_detected": "; ".join(detected_sequence_fields[:40]),
        "usable_for_training": "maybe",
        "usable_for_extension": "maybe",
        "priority": "high",
        "main_risk": (
            "Measured fluorescence/activity fields and predicted model-score fields coexist; "
            "standalone tables duplicate tables in the combined source-data workbook; units require paper/manual confirmation."
        ),
        "notes": (
            "Primary training candidate is diagnostic_activity. Do not merge predicted scores, figure summaries, "
            "or metadata with measured activity labels. source_id=EasyDesign_2024"
        ),
    }
    combined = pd.concat([existing, pd.DataFrame([row])], ignore_index=True)
    write_xlsx(combined, path, CATALOG_COLUMNS)
    return path


def normalize_label(raw_name: str) -> tuple[str, str, str, str, str, str, str, str]:
    lower = raw_name.lower()
    if lower in {"guide-activity", "mean-activity"}:
        return (
            "predicted_or_derived_activity_score",
            "predicted_library",
            "模型或覆盖分析产生的guide活性分数，不是直接实验标签",
            "Guide activity score produced by model or coverage analysis; not a direct experimental label",
            "model/derived activity score",
            "predicted",
            "no",
            "no",
        )
    if "30 min" in lower:
        return (
            "fluorescence_activity_30min",
            "diagnostic_activity",
            "30分钟Cas12a诊断荧光/活性读数",
            "Cas12a diagnostic fluorescence/activity readout at 30 minutes",
            "fluorescence/log activity",
            "measured",
            "maybe",
            "maybe",
        )
    if "20 min" in lower:
        return (
            "fluorescence_activity_20min_normalized",
            "diagnostic_activity",
            "20分钟归一化Cas12a诊断荧光/活性读数",
            "Normalized Cas12a diagnostic fluorescence/activity readout at 20 minutes",
            "normalized fluorescence/log activity",
            "measured",
            "maybe",
            "maybe",
        )
    if "out_logk" in lower or "logk" in lower:
        return (
            "activity_logk_measurement",
            "diagnostic_activity",
            "CRISPR荧光反应动力学相关logK测量值",
            "CRISPR fluorescence kinetic/logK-like measurement",
            "logK-like activity measurement",
            "measured",
            "maybe",
            "maybe",
        )
    if "true value" in lower:
        return (
            "measured_activity_true_value",
            "diagnostic_activity",
            "测试集实验真值/活性标签",
            "Experimental true value or activity label in the test dataset",
            "measured activity value",
            "measured",
            "yes",
            "maybe",
        )
    if re.search(r"fluoresc|fluoresence|rfu", lower):
        return (
            "fluorescence_raw_or_normalized",
            "diagnostic_activity",
            "Cas12a检测荧光读数，可能为原始或归一化值",
            "Cas12a diagnostic fluorescence readout, raw or normalized depending on figure/table context",
            "fluorescence/RFU",
            "measured",
            "maybe",
            "yes",
        )
    if re.search(r"cnnd|cnn12a|cnn12ae|transformer|expected|predicted|prediction", lower):
        return (
            "predicted_activity_score",
            "predicted_library",
            "模型预测活性分数，不是实验标签",
            "Model-predicted activity score; not an experimental label",
            "model prediction score",
            "predicted",
            "no",
            "no",
        )
    if re.search(r"spearman|pearson|coefficient|fold", lower):
        return (
            "model_performance_metric",
            "metadata_only",
            "模型性能评价指标，不是样本级标签",
            "Model performance metric, not a sample-level label",
            "correlation/performance metric",
            "metadata",
            "no",
            "no",
        )
    return (
        "unclear_label_candidate",
        "unclear",
        "疑似标签列，但含义需要人工确认",
        "Possible label column; meaning requires manual confirmation",
        "unclear",
        "unclear",
        "no",
        "maybe",
    )


def update_label_dictionary(audits: list[SheetAudit], catalog_dir: Path, timestamp: str) -> Path:
    path = catalog_dir / "label_dictionary.xlsx"
    backup_if_exists(path, timestamp)
    existing = read_existing_xlsx(path, LABEL_COLUMNS)
    if "notes" in existing.columns:
        existing = existing[~existing["notes"].str.contains("source_id=EasyDesign_2024", na=False)]
    labels: dict[str, dict[str, str]] = {}
    discovered = []
    for audit in audits:
        for raw_name in audit.label_raw_candidates:
            discovered.append((raw_name, audit.path_type_guess, audit.sheet_name, audit.file_name))
    discovered.extend(
        [
            ("guide-expected-activities", "predicted_library", "README output field", "README.md"),
            ("guide-activity", "predicted_library", "code output field", "bin/analyze_coverage.py"),
            ("mean-activity", "predicted_library", "code output field", "bin/analyze_coverage.py"),
        ]
    )
    for raw_name, path_type_hint, sheet_name, file_name in discovered:
        normalized, path_type, meaning_cn, meaning_en, assay, status, trainable, transform = normalize_label(raw_name)
        if path_type_hint in {
            "diagnostic_activity",
            "predicted_library",
            "metadata_only",
            "snv_specificity",
            "snv_annotation",
        } and path_type == "unclear":
            path_type = path_type_hint
        key = (raw_name, path_type)
        note = f"source_id=EasyDesign_2024; observed in {file_name} / {sheet_name}"
        labels["||".join(key)] = {
            "label_raw_name": raw_name,
            "normalized_label": normalized,
            "path_type": path_type,
            "biological_meaning_cn": meaning_cn,
            "biological_meaning_en": meaning_en,
            "assay_readout": assay,
            "label_status": status,
            "trainable_as_primary_label": trainable,
            "standard_unit": "unknown_from_source_table",
            "transform_needed": transform,
            "notes": note,
        }
    new_df = pd.DataFrame(list(labels.values()), columns=LABEL_COLUMNS)
    combined = pd.concat([existing, new_df], ignore_index=True)
    write_xlsx(combined, path, LABEL_COLUMNS)
    return path


def read_sheet_dataframe(file_path: Path, sheet_name: str, header_row: int) -> pd.DataFrame:
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row - 1, engine="openpyxl")
    df = df.dropna(how="all")
    df.columns = [clean_cell(col, max_len=120) or f"column_{idx+1}" for idx, col in enumerate(df.columns)]
    return df


def export_high_priority_csvs(audits: list[SheetAudit], root: Path, out_dir: Path) -> list[dict[str, str]]:
    exported = []
    for audit in audits:
        if not (
            audit.priority == "high"
            and audit.path_type_guess == "diagnostic_activity"
            and audit.recommended_action == "extract_to_csv"
        ):
            continue
        workbook_path = root / audit.file_path
        if workbook_path.name.lower().startswith("imt2214-sup"):
            table_id = sanitize_id(audit.sheet_name)
        else:
            table_id = sanitize_id(f"{workbook_path.stem}_{audit.sheet_name}")
        out_name = f"EasyDesign_2024_{table_id}_diagnostic_activity_raw.csv"
        out_path = out_dir / out_name
        df = read_sheet_dataframe(workbook_path, audit.sheet_name, audit.header_row)
        df.to_csv(out_path, index=False, encoding="utf-8-sig")
        audit.extracted_csv_path = safe_rel(out_path, root)
        exported.append(
            {
                "sheet": f"{audit.file_name} / {audit.sheet_name}",
                "csv_path": safe_rel(out_path, root),
                "rows": str(len(df)),
                "cols": str(len(df.columns)),
            }
        )
    return exported


def basic_quality_checks(audits: list[SheetAudit], root: Path) -> list[dict[str, str]]:
    checks = []
    seq_col_re = re.compile(r"guide_seq|target_at_guide|dna|crrna|sequence", re.IGNORECASE)
    valid_seq_re = re.compile(r"^[ACGTUNacgtun\s-]+$")
    for audit in audits:
        if audit.priority != "high" or audit.path_type_guess != "diagnostic_activity":
            continue
        df = read_sheet_dataframe(root / audit.file_path, audit.sheet_name, audit.header_row)
        seq_cols = [col for col in df.columns if seq_col_re.search(str(col))]
        invalid_counts = {}
        for col in seq_cols:
            series = df[col].dropna().astype(str).str.strip()
            checked = series[series != ""]
            invalid_counts[col] = int((~checked.str.match(valid_seq_re)).sum())
        checks.append(
            {
                "sheet": f"{audit.file_name} / {audit.sheet_name}",
                "rows": str(len(df)),
                "cols": str(len(df.columns)),
                "duplicate_rows": str(int(df.duplicated().sum())),
                "sequence_columns": "; ".join(seq_cols),
                "invalid_sequence_like_values": json.dumps(invalid_counts, ensure_ascii=False),
            }
        )
    return checks


def md_table(rows: list[dict[str, Any]], columns: list[str], max_rows: int | None = None) -> str:
    if max_rows is not None:
        rows = rows[:max_rows]
    if not rows:
        return "_None._\n"
    lines = []
    lines.append("| " + " | ".join(columns) + " |")
    lines.append("| " + " | ".join(["---"] * len(columns)) + " |")
    for row in rows:
        vals = []
        for col in columns:
            val = clean_cell(row.get(col, ""), max_len=300)
            val = val.replace("|", "\\|")
            vals.append(val)
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines) + "\n"


def audit_summary_counts(audits: list[SheetAudit]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for audit in audits:
        key = f"{audit.path_type_guess}:{audit.priority}"
        counts[key] = counts.get(key, 0) + 1
    return counts


def unresolved_questions(audits: list[SheetAudit]) -> list[str]:
    questions = [
        "确认 `30 min` 与 `20 min normalized` 的准确单位和转换关系：它们是否均为实验荧光/活性读数，还是一个为归一化派生标签？",
        "确认 `out_logk_measurement` 是否由同一实验读数转换而来，是否可作为 primary diagnostic activity label。",
        "确认 `true value` 在 Test data/Table S5 中的来源：是否为实验测量值，以及是否与训练集标签处在同一尺度。",
        "确认 standalone workbooks (`Table S1.xlsx`-`Table S4.xlsx`) 与 combined source-data workbook (`imt2214-sup-0002...xlsx`) 的重复关系，后续清洗时只保留哪一套来源。",
        "Figure source data 中的 normalized fluorescence/fluorescence unit 是否能映射回具体 crRNA-target pair；当前不能直接作为 tidy training rows。",
        "README/代码中的 `guide-expected-activities` 是模型预测输出，不能当作实验标签；若后续使用，需要明确作为 predicted_library 或辅助特征。",
        "未在当前 Excel 结构中发现明确 PAM 列；需要确认 Cas12a PAM 是否隐含在 target/context sequence 中。",
    ]
    for audit in audits:
        if audit.path_type_guess == "unclear" or audit.has_label == "unclear":
            questions.append(f"人工确认 {audit.file_name} / {audit.sheet_name} 的内容类型和标签含义。")
    return questions


def write_sectioned_file(path: Path, title: str, body: str) -> None:
    path.write_text(f"# {title}\n\n{body.rstrip()}\n", encoding="utf-8")


def replace_marked_section(path: Path, marker: str, content: str) -> None:
    start = f"<!-- BEGIN {marker} -->"
    end = f"<!-- END {marker} -->"
    section = f"{start}\n{content.rstrip()}\n{end}\n"
    existing = path.read_text(encoding="utf-8", errors="replace") if path.exists() else ""
    pattern = re.compile(re.escape(start) + r".*?" + re.escape(end) + r"\n?", re.DOTALL)
    if pattern.search(existing):
        new_text = pattern.sub(section, existing)
    else:
        prefix = existing.rstrip() + "\n\n" if existing.strip() else ""
        new_text = prefix + section
    path.write_text(new_text, encoding="utf-8")


def write_reports(
    root: Path,
    paths: dict[str, Path],
    inventory: list[dict[str, str]],
    doc_paths: list[str],
    doc_facts: list[dict[str, str]],
    audits: list[SheetAudit],
    exported: list[dict[str, str]],
    quality: list[dict[str, str]],
    outputs: list[str],
) -> dict[str, Path]:
    notes_dir = paths["notes"]
    counts = audit_summary_counts(audits)
    sheet_rows = [audit.as_row() for audit in audits]
    high_rows = [
        audit.as_row()
        for audit in audits
        if audit.priority == "high" and audit.path_type_guess == "diagnostic_activity"
    ]
    label_rows = []
    for audit in audits:
        for label in audit.label_raw_candidates:
            _normalized, label_path_type, _cn, _en, _assay, label_status, _trainable, _transform = normalize_label(label)
            label_rows.append(
                {
                    "file_sheet": f"{audit.file_name} / {audit.sheet_name}",
                    "label_candidate": label,
                    "status": label_status,
                    "path_type": label_path_type if label_path_type != "unclear" else audit.path_type_guess,
                }
            )
    unresolved = unresolved_questions(audits)

    audit_body = "\n".join(
        [
            "## Scope",
            f"Raw source audited: `01_raw/{SOURCE_ID}`. No files under `01_raw` were modified.",
            "",
            "## Directory And File Inventory",
            md_table(inventory, ["file_path", "file_type", "file_size", "possible_usage"]),
            "## Documentation Found",
            "Documentation/config files scanned: " + (", ".join(f"`{p}`" for p in doc_paths) if doc_paths else "未找到 README 或说明文件"),
            "",
            "### Confirmed Facts From Documentation",
            md_table(doc_facts[:20], ["file_path", "line", "evidence"]) if doc_facts else "未找到 README 或说明文件。\n",
            "## Workbook Sheet Audit",
            md_table(
                sheet_rows,
                [
                    "file_name",
                    "sheet_name",
                    "n_rows",
                    "n_cols",
                    "path_type_guess",
                    "record_unit_guess",
                    "has_crRNA_sequence",
                    "has_target_sequence",
                    "has_label",
                    "label_raw_candidates",
                    "priority",
                    "recommended_action",
                    "extracted_csv_path",
                ],
            ),
            "## Highest-Priority Sheets For Later Cleaning",
            md_table(
                high_rows,
                [
                    "file_name",
                    "sheet_name",
                    "n_rows",
                    "n_cols",
                    "label_raw_candidates",
                    "extracted_csv_path",
                    "notes",
                ],
            ),
            "## Possible Label Fields",
            md_table(label_rows, ["file_sheet", "label_candidate", "status", "path_type"]),
            "## Key Findings",
            "- Confirmed from README: EasyDesign is a Cas12a diagnostic crRNA design system and its output includes guide expected activities, target sequences, and target positions.",
            "- Confirmed from workbooks: `Training data`, `Augment data`, and `Test data` plus combined workbook `Table S3/S4/S5` contain guide/crRNA plus target/DNA sequence fields and activity-like measured labels.",
            "- Initial inference: figure source-data sheets contain fluorescence or activity-like values but are not tidy crRNA-target pair tables, so they were not exported as high-priority training candidates.",
            "- Initial inference: model columns such as `CNND`, `CNN12a`, and `Transformer12a*` are predicted scores and must not be used as primary measured labels.",
            "",
            "## Unresolved Questions",
            "\n".join(f"- {q}" for q in unresolved),
        ]
    )
    audit_path = notes_dir / "EasyDesign_2024_data_audit.md"
    write_sectioned_file(audit_path, "EasyDesign_2024 Data Audit", audit_body)

    run_body = "\n".join(
        [
            "## Scope",
            f"本次处理范围为 `01_raw/{SOURCE_ID}`。遵守只读 raw 文件原则；未训练模型，未合并不同标签体系。",
            "",
            "## Inputs Scanned",
            md_table(inventory, ["file_path", "file_type", "file_size", "possible_usage"]),
            "## Outputs Generated",
            "\n".join(f"- `{path}`" for path in outputs),
            "",
            "## Classification Decisions",
            "Summary counts: " + json.dumps(counts, ensure_ascii=False),
            "",
            md_table(
                sheet_rows,
                [
                    "file_name",
                    "sheet_name",
                    "guessed_content",
                    "record_unit_guess",
                    "label_status_guess",
                    "path_type_guess",
                    "priority",
                    "recommended_action",
                ],
            ),
            "## Evidence",
            "### Confirmed Facts",
            "- README states that EasyDesign is an enhanced crRNA design system using deep learning for Cas12a-mediated diagnostics.",
            "- README states main output fields include `guide-expected-activities`, `guide-target-sequences`, and `guide-target-sequence-positions`.",
            "- Excel sheets were scanned programmatically for sheet name, dimensions, columns, and representative values.",
            "",
            "### Initial Inferences From Sheet Names And Columns",
            "- Sheets with `guide_seq`/`target_at_guide` plus `30 min`, `20 min normalized`, `out_logk_measurement`, or `true value` were classified as diagnostic activity candidates.",
            "- Sheets with `CNND`, `CNN12a`, `Transformer*`, `expected`, or `predicted` were treated as containing model-derived scores, not primary measured labels.",
            "- Figure sheets with fluorescence values but without tidy sequence fields were marked for manual inspection rather than direct extraction.",
            "",
            "## Data Quality Checks",
            md_table(quality, ["sheet", "rows", "cols", "duplicate_rows", "sequence_columns", "invalid_sequence_like_values"]),
            "## Extracted Raw CSVs",
            md_table(exported, ["sheet", "csv_path", "rows", "cols"]),
            "## Unresolved Questions",
            "\n".join(f"- {q}" for q in unresolved),
            "",
            "## Next Recommended Actions",
            "- Manually confirm label units for `30 min`, `20 min normalized`, `out_logk_measurement`, and `true value` before any minimal cleaning.",
            "- Decide whether to use standalone `Table S2.xlsx` sheets or combined source-data workbook `Table S3/S4/S5` as the authoritative extraction source to avoid duplicates.",
            "- After confirmation, build `03_cleaned_minimal/diagnostic_activity_minimal.csv` with traceable source_table_id and raw label fields; do not proceed directly to model training.",
        ]
    )
    run_path = notes_dir / "run_report_EasyDesign_2024.md"
    write_sectioned_file(run_path, "Run Report EasyDesign_2024", run_body)

    evidence_lines = []
    for audit in audits:
        evidence_lines.extend(
            [
                f"## {audit.file_name} / {audit.sheet_name}",
                "",
                f"Decision: `{audit.path_type_guess}` with priority `{audit.priority}`.",
                "",
                "Evidence:",
                f"- File name: `{audit.file_name}`",
                f"- Sheet name: `{audit.sheet_name}`",
                f"- Title/caption: {audit.title_or_caption or 'none detected'}",
                f"- First columns: {', '.join(audit.first_columns[:20])}",
                f"- Label candidates: {', '.join(audit.label_raw_candidates) or 'none detected'}",
                "",
                "Reasoning:",
                f"- {audit.guessed_content}",
                "",
                "Remaining uncertainty:",
                f"- {audit.notes or 'No specific uncertainty beyond normal label-unit confirmation.'}",
                "",
            ]
        )
    evidence_path = notes_dir / "evidence_trace_EasyDesign_2024.md"
    write_sectioned_file(evidence_path, "Evidence Trace EasyDesign_2024", "\n".join(evidence_lines))

    problems_path = notes_dir / "problems_to_resolve.md"
    problems_content = "\n".join(
        [
            "## EasyDesign_2024",
            "",
            "\n".join(f"- {q}" for q in unresolved),
        ]
    )
    replace_marked_section(problems_path, "EasyDesign_2024", problems_content)

    method_path = notes_dir / "method_notes.md"
    method_content = (
        "## EasyDesign_2024 Data Audit Method\n\n"
        "Raw supplementary files were stored without modification. Excel workbooks were "
        "programmatically scanned using Python, pandas, and openpyxl. Each worksheet was "
        "indexed by file name, sheet name, dimensions, detected header row, first columns, "
        "and representative values. Sheets were classified by data type based on detected "
        "sequence fields, candidate label fields, source documentation, sheet names, and "
        "table captions. Experimentally measured diagnostic activity candidates were "
        "exported only as raw traceable CSV tables under `02_extracted_tables/diagnostic_activity/`; "
        "no final training dataset was built and no model was trained. Predicted activity "
        "scores and model outputs were recorded separately from measured labels."
    )
    replace_marked_section(method_path, "EasyDesign_2024", method_content)

    return {
        "audit": audit_path,
        "run_report": run_path,
        "evidence": evidence_path,
        "problems": problems_path,
        "method": method_path,
    }


def main() -> None:
    root = project_root_from_script()
    paths = ensure_dirs(root)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    inventory = inventory_files(paths["raw"], root)
    doc_facts, doc_paths = extract_doc_facts(paths["raw"], root)
    audits = scan_workbooks(paths["raw"], root)

    exported = export_high_priority_csvs(audits, root, paths["diagnostic"])
    quality = basic_quality_checks(audits, root)

    sheet_index_path = update_sheet_index(audits, paths["catalog"], timestamp)
    master_path = update_master_catalog(audits, inventory, paths["catalog"], timestamp)
    label_path = update_label_dictionary(audits, paths["catalog"], timestamp)

    outputs = [
        safe_rel(sheet_index_path, root),
        safe_rel(master_path, root),
        safe_rel(label_path, root),
    ]
    outputs.extend(row["csv_path"] for row in exported)

    report_paths = write_reports(
        root=root,
        paths=paths,
        inventory=inventory,
        doc_paths=doc_paths,
        doc_facts=doc_facts,
        audits=audits,
        exported=exported,
        quality=quality,
        outputs=outputs,
    )
    outputs.extend(safe_rel(path, root) for path in report_paths.values())

    print(json.dumps({"source_id": SOURCE_ID, "outputs": outputs}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
